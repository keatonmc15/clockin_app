import os
import math
import logging
import csv
import json
import re
import threading
import time
import hmac
from io import TextIOWrapper
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, timedelta, time as dtime

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, Response
)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, text, select
from sqlalchemy.exc import IntegrityError

# ✅ XLSX export support
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor, XDRPositiveSize2D
from openpyxl.styles import Font, Alignment, Border, PatternFill, Side
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils import get_column_letter

# -----------------------------
# Timezone (Windows-safe)
# -----------------------------
APP_TZ = None
UTC_TZ = None
try:
    from zoneinfo import ZoneInfo
    try:
        APP_TZ = ZoneInfo("America/Chicago")
    except Exception:
        APP_TZ = None
    try:
        UTC_TZ = ZoneInfo("UTC")
    except Exception:
        UTC_TZ = None
except Exception:
    APP_TZ = None
    UTC_TZ = None

# -----------------------------
# App + Config
# -----------------------------
app = Flask(__name__)

# Basic INFO logging (Render captures these)
logging.basicConfig(level=logging.INFO)

def _required_env(name: str) -> str:
    value = (os.environ.get(name) or "").strip()
    if not value:
        raise RuntimeError(f"{name} environment variable is required")
    return value

app.config["SECRET_KEY"] = _required_env("SECRET_KEY")

# ✅ Mobile ingest auth token (set on Render)
app.config["MOBILE_DEVICE_TOKEN"] = _required_env("MOBILE_DEVICE_TOKEN")

# ✅ Dev endpoint gate
ENABLE_DEV_EXPORTS = (os.environ.get("ENABLE_DEV_EXPORTS") or "").strip() == "1"

# ------------------------------------------------------------
# Database config
#   - Prefer DATABASE_URL if present (Render)
#   - Still supports USE_RENDER_DB=1 if you like
#   - Local fallback: sqlite relative path
# ------------------------------------------------------------
def _normalize_db_url(raw: str) -> str:
    raw = (raw or "").strip()
    if not raw:
        return raw

    # Normalize Render postgres URL
    if raw.startswith("postgres://"):
        raw = raw.replace("postgres://", "postgresql://", 1)

    # FORCE psycopg v3 (required for Python 3.13)
    if raw.startswith("postgresql://") and not raw.startswith("postgresql+psycopg://"):
        raw = raw.replace("postgresql://", "postgresql+psycopg://", 1)

    return raw


db_url = _normalize_db_url(_required_env("DATABASE_URL"))

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

AUTO_EXIT_GRACE_SECONDS = 5 * 60
AUTO_EXIT_ACCURACY_MAX_M = 120.0
AUTO_EXIT_OUTSIDE_BUFFER_M = 15.0
AUTO_EXIT_OUTSIDE_CONFIRM_SECONDS = 60
AUTO_EXIT_MAX_EVENT_AGE_SECONDS = 15 * 60
BG_LOCATION_PING_INTERVAL = timedelta(minutes=15)
AUTO_EXIT_WORKER_INTERVAL_SECONDS = 60

# -----------------------------
# Flask-Migrate (optional)
# -----------------------------
try:
    from flask_migrate import Migrate
    migrate = Migrate(app, db)
except Exception:
    migrate = None

# -----------------------------
# Admin credentials
# -----------------------------
ADMIN_USERNAME = _required_env("ADMIN_USERNAME")
ADMIN_PASSWORD = _required_env("ADMIN_PASSWORD")
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH") or generate_password_hash(ADMIN_PASSWORD)

# -----------------------------
# Helpers (time)
# -----------------------------
def now_utc() -> datetime:
    # store naive UTC (works reliably with db.DateTime columns)
    return datetime.utcnow()

def now_local() -> datetime:
    if APP_TZ:
        return datetime.now(APP_TZ)
    return datetime.now()

def utc_naive_to_local(dt: datetime) -> datetime:
    """
    Treat incoming dt as UTC if naive; convert to APP_TZ for display.
    """
    if not dt:
        return dt
    try:
        if getattr(dt, "tzinfo", None) is None:
            if UTC_TZ:
                dt = dt.replace(tzinfo=UTC_TZ)
        if APP_TZ and getattr(dt, "tzinfo", None):
            dt = dt.astimezone(APP_TZ)
    except Exception:
        pass
    return dt

# -----------------------------
# Models
# -----------------------------
class Store(db.Model):
    __tablename__ = "stores"
    id = db.Column(db.Integer, primary_key=True)

    name = db.Column(db.String(120), nullable=False)
    qr_token = db.Column(db.String(120), unique=True, nullable=False)

    latitude = db.Column(db.Float, nullable=False)
    longitude = db.Column(db.Float, nullable=False)
    geofence_radius_m = db.Column(db.Integer, nullable=False, default=150)

    created_at = db.Column(db.DateTime, default=lambda: now_utc())

class Employee(db.Model):
    __tablename__ = "employees"
    id = db.Column(db.Integer, primary_key=True)

    name = db.Column(db.String(120), nullable=False)
    username_code = db.Column(db.String(80), unique=True, nullable=True)
    pin = db.Column(db.String(20), nullable=False)
    active = db.Column(db.Boolean, default=True, nullable=False)

    # ✅ Option 2: device binding (Option C = overwrite allowed)
    device_uuid = db.Column(db.String(120), nullable=True)        # last seen/bound device
    device_label = db.Column(db.String(120), nullable=True)       # optional "Pixel 7", etc
    device_last_seen_at = db.Column(db.DateTime, nullable=True)   # UTC naive

    created_at = db.Column(db.DateTime, default=lambda: now_utc())

class Shift(db.Model):
    __tablename__ = "shifts"
    id = db.Column(db.Integer, primary_key=True)

    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey("stores.id"), nullable=False)

    clock_in = db.Column(db.DateTime, nullable=False)
    clock_out = db.Column(db.DateTime, nullable=True)

    clock_in_lat = db.Column(db.Float, nullable=True)
    clock_in_lng = db.Column(db.Float, nullable=True)
    clock_out_lat = db.Column(db.Float, nullable=True)
    clock_out_lng = db.Column(db.Float, nullable=True)

    # ✅ Option 2: capture device uuid on punches
    clock_in_device_uuid = db.Column(db.String(120), nullable=True)
    clock_out_device_uuid = db.Column(db.String(120), nullable=True)
    clock_out_source = db.Column(db.String(32), nullable=True)

    # --- Admin override audit fields (B) ---
    closed_by_admin = db.Column(db.Boolean, nullable=False, default=False)
    admin_closed_by = db.Column(db.String(120), nullable=True)   # username
    admin_closed_at = db.Column(db.DateTime, nullable=True)
    admin_close_reason = db.Column(db.Text, nullable=True)

    created_at = db.Column(db.DateTime, default=lambda: now_utc())

    employee = db.relationship("Employee", backref="shifts")
    store = db.relationship("Store", backref="shifts")

# ✅ Location pings (15-min tracking)
class LocationPing(db.Model):
    __tablename__ = "location_pings"
    id = db.Column(db.Integer, primary_key=True)

    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    shift_id = db.Column(db.Integer, db.ForeignKey("shifts.id"), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey("stores.id"), nullable=False)

    lat = db.Column(db.Float, nullable=False)
    lng = db.Column(db.Float, nullable=False)

    dist_m = db.Column(db.Float, nullable=False)
    inside_radius = db.Column(db.Boolean, nullable=False, default=True)

    created_at = db.Column(db.DateTime, default=lambda: now_utc(), nullable=False)

    employee = db.relationship("Employee")
    shift = db.relationship("Shift")
    store = db.relationship("Store")

# ✅ NEW: Shift edit audit trail (Option B-safe: new table)
class ShiftEditAudit(db.Model):
    __tablename__ = "shift_edit_audit"
    id = db.Column(db.Integer, primary_key=True)

    shift_id = db.Column(db.Integer, db.ForeignKey("shifts.id"), nullable=True)
    action = db.Column(db.String(40), nullable=False)  # create/edit/force_close
    editor = db.Column(db.String(120), nullable=False)  # admin username
    reason = db.Column(db.Text, nullable=False)

    old_clock_in = db.Column(db.DateTime, nullable=True)
    old_clock_out = db.Column(db.DateTime, nullable=True)
    new_clock_in = db.Column(db.DateTime, nullable=True)
    new_clock_out = db.Column(db.DateTime, nullable=True)

    created_at = db.Column(db.DateTime, default=lambda: now_utc(), nullable=False)

    shift = db.relationship("Shift")

# ✅ Mobile ingest raw event store (Option B-safe: new table)
class MobileEvent(db.Model):
    __tablename__ = "mobile_events"
    id = db.Column(db.Integer, primary_key=True)

    event_type = db.Column(db.String(50), nullable=False, default="unknown")
    device_uuid = db.Column(db.String(120), nullable=True)
    is_moving = db.Column(db.Boolean, nullable=True)

    lat = db.Column(db.Float, nullable=True)
    lng = db.Column(db.Float, nullable=True)
    accuracy = db.Column(db.Float, nullable=True)

    event_at = db.Column(db.DateTime, nullable=True)
    received_at = db.Column(db.DateTime, default=lambda: now_utc(), nullable=False)

    raw_json = db.Column(db.Text, nullable=False)

class PendingAutoExit(db.Model):
    __tablename__ = "pending_auto_exits"
    __table_args__ = (
        db.Index("ix_pending_auto_exits_shift_status", "shift_id", "status"),
        db.Index("ix_pending_auto_exits_status_deadline", "status", "deadline_at"),
        db.Index("ix_pending_auto_exits_employee_status", "employee_id", "status"),
        db.Index("ix_pending_auto_exits_store_status", "store_id", "status"),
        db.Index(
            "ux_pending_auto_exits_open_shift",
            "shift_id",
            unique=True,
            postgresql_where=text("status IN ('active', 'candidate')"),
            sqlite_where=text("status IN ('active', 'candidate')"),
        ),
    )
    id = db.Column(db.Integer, primary_key=True)

    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=False)
    shift_id = db.Column(db.Integer, db.ForeignKey("shifts.id"), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey("stores.id"), nullable=False)

    device_uuid = db.Column(db.String(120), nullable=True)
    status = db.Column(db.String(30), nullable=False, default="active")
    source = db.Column(db.String(50), nullable=False, default="outside_location")

    exit_at = db.Column(db.DateTime, nullable=False, default=lambda: now_utc())
    deadline_at = db.Column(db.DateTime, nullable=False)
    last_seen_at = db.Column(db.DateTime, nullable=True)

    last_lat = db.Column(db.Float, nullable=True)
    last_lng = db.Column(db.Float, nullable=True)
    last_accuracy_m = db.Column(db.Float, nullable=True)
    last_dist_m = db.Column(db.Float, nullable=True)

    outside_count = db.Column(db.Integer, nullable=False, default=1)
    cancel_reason = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, default=lambda: now_utc(), nullable=False)
    updated_at = db.Column(db.DateTime, default=lambda: now_utc(), nullable=False)
    closed_at = db.Column(db.DateTime, nullable=True)

    employee = db.relationship("Employee")
    shift = db.relationship("Shift")
    store = db.relationship("Store")

class MobileIssueReport(db.Model):
    __tablename__ = "mobile_issue_reports"
    id = db.Column(db.Integer, primary_key=True)

    employee_id = db.Column(db.Integer, db.ForeignKey("employees.id"), nullable=True)
    store_id = db.Column(db.Integer, db.ForeignKey("stores.id"), nullable=True)
    shift_id = db.Column(db.Integer, db.ForeignKey("shifts.id"), nullable=True)

    message = db.Column(db.Text, nullable=True)
    payload_json = db.Column(db.Text, nullable=False, default="{}")

    status = db.Column(db.String(30), nullable=False, default="open")  # open / resolved / ignored
    resolved_by = db.Column(db.String(120), nullable=True)
    resolved_at = db.Column(db.DateTime, nullable=True)
    resolve_note = db.Column(db.Text, nullable=True)

    created_at = db.Column(db.DateTime, default=lambda: now_utc(), nullable=False)

    employee = db.relationship("Employee")
    store = db.relationship("Store")
    shift = db.relationship("Shift")

# -----------------------------
# Geo Helpers
# -----------------------------
def haversine_m(lat1, lon1, lat2, lon2) -> float:
    """
    Returns distance in meters between two WGS84 lat/lon points.
    """
    R = 6371000.0  # meters
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)

    a = math.sin(dphi / 2.0) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlmb / 2.0) ** 2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return R * c

def find_store_for_location(
    lat: float,
    lon: float,
    accuracy_m: float | None = None,
    *,
    max_accuracy_m: float = 120.0,
    sanity_gap_m: float = 800.0,
):
    """
    Returns a dict with the best-matching store and distance, or None if not inside any store geofence.
    """
    if accuracy_m is not None and accuracy_m > max_accuracy_m:
        return {
            "ok": False,
            "reason": "accuracy_too_low",
            "message": "GPS accuracy is too low. Step outside and try again.",
            "accuracy_m": float(accuracy_m),
            "max_accuracy_m": float(max_accuracy_m),
        }

    stores = db.session.execute(select(Store)).scalars().all()
    if not stores:
        return {"ok": False, "reason": "no_stores", "message": "No stores are configured."}

    distances = []
    for s in stores:
        d = haversine_m(lat, lon, s.latitude, s.longitude)
        distances.append((d, s))

    distances.sort(key=lambda x: x[0])
    best_d, best_store = distances[0]

    if len(distances) > 1:
        second_d, _ = distances[1]
        if (second_d - best_d) < sanity_gap_m:
            return {
                "ok": False,
                "reason": "ambiguous_nearest",
                "message": "Location is ambiguous between two stores. Move closer to the building and try again.",
                "best_distance_m": float(best_d),
                "second_distance_m": float(second_d),
                "sanity_gap_m": float(sanity_gap_m),
            }

    if best_d <= best_store.geofence_radius_m:
        return {
            "ok": True,
            "store": best_store,
            "distance_m": float(best_d),
        }

    return {
        "ok": False,
        "reason": "outside_geofence",
        "message": "You are not within a valid store location.",
        "nearest_store_id": int(best_store.id),
        "nearest_store_name": best_store.name,
        "nearest_distance_m": float(best_d),
        "required_radius_m": float(best_store.geofence_radius_m),
    }

# -----------------------------
# Helpers (general)
# -----------------------------
def fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return ""
    dt_local = utc_naive_to_local(dt)
    return dt_local.strftime("%Y-%m-%d %I:%M %p")

def parse_local_datetime(val: str) -> datetime | None:
    """
    Accepts 'YYYY-MM-DDTHH:MM' (HTML datetime-local) OR 'YYYY-MM-DD HH:MM'
    Input is interpreted as America/Chicago (APP_TZ) and converted to UTC-naive for storage.
    """
    if not val:
        return None
    s = val.strip()
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
        try:
            naive = datetime.strptime(s, fmt)  # naive local wall time
            if APP_TZ and UTC_TZ:
                local_dt = naive.replace(tzinfo=APP_TZ)
                utc_dt = local_dt.astimezone(UTC_TZ)
                return utc_dt.replace(tzinfo=None)  # store naive UTC
            return naive
        except ValueError:
            continue
    return None

def local_range_to_utc_naive(start_local: datetime, end_local: datetime) -> tuple[datetime, datetime]:
    """
    Converts tz-aware local bounds (America/Chicago) to UTC-naive for DB filtering.
    """
    if APP_TZ and UTC_TZ and getattr(start_local, "tzinfo", None) and getattr(end_local, "tzinfo", None):
        s_utc = start_local.astimezone(UTC_TZ).replace(tzinfo=None)
        e_utc = end_local.astimezone(UTC_TZ).replace(tzinfo=None)
        return s_utc, e_utc
    return start_local.replace(tzinfo=None), end_local.replace(tzinfo=None)

# ✅ Step 1: minute-accurate shift minutes (NO quarter-hour rounding)
def shift_minutes(shift: "Shift") -> int:
    if not shift.clock_in or not shift.clock_out:
        return 0
    seconds = (shift.clock_out - shift.clock_in).total_seconds()
    if seconds <= 0:
        return 0
    return int(seconds // 60)  # whole minutes

def minutes_to_human(minutes: int) -> str:
    if minutes <= 0:
        return "0 min"
    hours = minutes // 60
    mins = minutes % 60
    if hours > 0 and mins > 0:
        return f"{hours} hr {mins} min"
    elif hours > 0:
        return f"{hours} hr"
    else:
        return f"{mins} min"

def minutes_to_short(minutes: int) -> str:
    if minutes <= 0:
        return "0h 00m"
    h = minutes // 60
    m = minutes % 60
    return f"{h}h {m:02d}m"

def minutes_to_decimal_hours(minutes: int, places: int = 4) -> str:
    if minutes <= 0:
        return "0"
    val = (Decimal(minutes) / Decimal(60)).quantize(
        Decimal("1." + "0" * places),
        rounding=ROUND_HALF_UP
    )
    return format(val, "f")

def shift_hours(shift: "Shift") -> float:
    mins = shift_minutes(shift)
    return float(Decimal(mins) / Decimal(60)) if mins else 0.0

def shift_clock_out_source(shift: "Shift") -> str:
    if not shift or not shift.clock_out:
        return "open"
    source = (getattr(shift, "clock_out_source", None) or "").strip().lower()
    if source in {"employee", "auto_exit", "admin"}:
        return source
    if (getattr(shift, "admin_closed_by", None) or "").strip().upper() == "AUTO_EXIT":
        return "auto_exit"
    if getattr(shift, "closed_by_admin", False):
        return "admin"
    return "employee"

def shift_clock_out_label(shift: "Shift") -> str:
    return {
        "employee": "Employee Clock-Out",
        "auto_exit": "Auto Exit",
        "admin": "Admin Close",
        "open": "Open",
    }.get(shift_clock_out_source(shift), "Employee Clock-Out")

def shift_clock_out_distance_m(shift: "Shift") -> float | None:
    if not shift or shift.clock_out_lat is None or shift.clock_out_lng is None or not shift.store:
        return None
    try:
        return float(haversine_m(shift.clock_out_lat, shift.clock_out_lng, shift.store.latitude, shift.store.longitude))
    except Exception:
        return None

def shift_clock_out_outside_geofence(shift: "Shift") -> bool:
    dist = shift_clock_out_distance_m(shift)
    if dist is None or not shift or not shift.store:
        return False
    return dist > shift.store.geofence_radius_m

def active_pending_auto_exit_for_shift(shift: "Shift"):
    if not shift or shift.clock_out:
        return None
    return (
        PendingAutoExit.query
        .filter(PendingAutoExit.shift_id == shift.id, PendingAutoExit.status == "active")
        .order_by(PendingAutoExit.deadline_at.asc())
        .first()
    )

def pending_auto_exit_minutes_remaining(pending: "PendingAutoExit") -> int:
    if not pending or not pending.deadline_at:
        return 0
    remaining = (pending.deadline_at - now_utc()).total_seconds()
    return max(0, int(math.ceil(remaining / 60.0)))

def last_completed_payroll_week(reference: datetime | None = None):
    ref_local = reference or now_local()
    weekday = ref_local.weekday()  # Monday=0
    this_monday = ref_local.date() - timedelta(days=weekday)
    last_monday = this_monday - timedelta(days=7)
    last_sunday = last_monday + timedelta(days=6)

    if APP_TZ:
        start_dt = datetime.combine(last_monday, dtime.min, tzinfo=APP_TZ)
        end_dt = datetime.combine(last_sunday, dtime.max, tzinfo=APP_TZ)
    else:
        start_dt = datetime.combine(last_monday, dtime.min)
        end_dt = datetime.combine(last_sunday, dtime.max)

    return start_dt, end_dt

def require_admin():
    return session.get("admin_logged_in") is True

def admin_guard():
    if not require_admin():
        return redirect(url_for("admin_login"))
    return None

def admin_username() -> str:
    return (session.get("admin_username") or ADMIN_USERNAME or "admin")

# ✅ Canonical store codes = lowercase
def normalize_store_code(val: str) -> str:
    return (val or "").strip().lower()

def normalize_employee_code(val: str) -> str:
    return (val or "").strip().lower()

def suggest_employee_username(name: str) -> str:
    parts = re.findall(r"[a-z0-9]+", (name or "").strip().lower())
    if len(parts) >= 2:
        base = f"{parts[0]}{parts[-1][0]}"
    elif parts:
        base = parts[0]
    else:
        base = ""
    return base or "employee"

def employee_code_exists(code: str, exclude_id: int | None = None) -> bool:
    code = normalize_employee_code(code)
    if not code:
        return False

    q = Employee.query.filter(func.lower(Employee.username_code) == code)
    if exclude_id is not None:
        q = q.filter(Employee.id != exclude_id)
    return q.first() is not None

def unique_employee_code_from_name(name: str, exclude_id: int | None = None) -> str:
    base = suggest_employee_username(name)
    candidate = base
    suffix = 2
    while employee_code_exists(candidate, exclude_id=exclude_id):
        candidate = f"{base}{suffix}"
        suffix += 1
    return candidate

def employee_payload(emp: "Employee") -> dict:
    return {
        "id": emp.id,
        "name": emp.name,
        "username_code": emp.username_code or "",
        "suggested_username_code": suggest_employee_username(emp.name),
        "active": bool(emp.active),
        "device_uuid": emp.device_uuid,
        "device_label": emp.device_label,
        "device_last_seen_at": fmt_dt(emp.device_last_seen_at) if emp.device_last_seen_at else "",
    }

def find_employee_for_mobile(username_code: str, pin: str) -> "Employee | None":
    code = normalize_employee_code(username_code)
    pin = (pin or "").strip()
    if not pin:
        return None

    if code:
        emp = (
            Employee.query
            .filter(func.lower(Employee.username_code) == code, Employee.pin == pin)
            .first()
        )
        if emp:
            return emp

        # Migration fallback: existing employees without a saved code may use
        # the generated suggestion shown in admin until the code is saved.
        legacy_matches = Employee.query.filter(Employee.pin == pin).all()
        for candidate in legacy_matches:
            if not (candidate.username_code or "").strip() and suggest_employee_username(candidate.name) == code:
                return candidate
        return None

    # Temporary legacy compatibility for old installed mobile builds.
    return Employee.query.filter_by(pin=pin).first()

def log_event(event: str, **fields):
    parts = [f"{k}={fields[k]}" for k in sorted(fields.keys())]
    app.logger.info("%s %s", event, " ".join(parts))

# -----------------------------
# Mobile ingest helpers
# -----------------------------
def _get_device_token() -> str:
    token = (request.headers.get("X-Device-Token") or "").strip()
    if token:
        return token

    auth = (request.headers.get("Authorization") or "").strip()
    if auth.lower().startswith("bearer "):
        return auth[7:].strip()

    return ""

def _require_mobile_auth():
    expected = (app.config.get("MOBILE_DEVICE_TOKEN") or "").strip()
    provided = _get_device_token()

    if not expected:
        # Fail closed: don't allow anonymous ingest if you forgot to set env var on Render
        app.logger.error("MOBILE_DEVICE_TOKEN is not set on the server.")
        return False, ("server_not_configured", 500)

    if not provided or not hmac.compare_digest(provided, expected):
        return False, ("unauthorized", 401)

    return True, None

def _dev_guard():
    if not ENABLE_DEV_EXPORTS:
        return False, ("not_found", 404)
    ok, err = _require_mobile_auth()
    if not ok:
        return False, err
    return True, None

def _safe_json_dumps(obj) -> str:
    try:
        return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    except Exception:
        return json.dumps({"_error": "json_dumps_failed"}, separators=(",", ":"))

def _extract_location_coords(payload: dict) -> tuple[dict, dict]:
    loc = {}
    if isinstance(payload.get("location"), dict):
        loc = payload.get("location") or {}
    elif isinstance(payload.get("geofence"), dict) and isinstance((payload.get("geofence") or {}).get("location"), dict):
        loc = (payload.get("geofence") or {}).get("location") or {}
    elif isinstance(payload.get("params"), dict) and isinstance((payload.get("params") or {}).get("location"), dict):
        loc = (payload.get("params") or {}).get("location") or {}
    elif isinstance(payload.get("data"), dict) and isinstance((payload.get("data") or {}).get("location"), dict):
        loc = (payload.get("data") or {}).get("location") or {}

    coords = (loc.get("coords") or {}) if isinstance(loc, dict) else {}
    if not isinstance(coords, dict):
        coords = {}
    if not coords and isinstance(payload.get("coords"), dict):
        coords = payload.get("coords") or {}
    return loc, coords

def _extract_event_at(payload: dict, loc: dict | None) -> datetime | None:
    ts_ms = payload.get("timestamp")
    if ts_ms is None and isinstance(loc, dict):
        ts_ms = loc.get("timestamp")
    if ts_ms is None and isinstance(payload.get("geofence"), dict):
        ts_ms = (payload.get("geofence") or {}).get("timestamp")

    if isinstance(ts_ms, (int, float)) and ts_ms > 0:
        try:
            return datetime.utcfromtimestamp(ts_ms / 1000.0)
        except Exception:
            return None
    return None

def _extract_device_uuid_from_bg_payload(payload: dict, loc: dict | None) -> str | None:
    candidates = [
        payload.get("device_uuid"),
        payload.get("uuid"),
    ]
    if isinstance(payload.get("device"), dict):
        candidates.append((payload.get("device") or {}).get("uuid"))
    if isinstance(loc, dict):
        candidates.append(loc.get("uuid"))
    if isinstance(payload.get("geofence"), dict):
        gf = payload.get("geofence") or {}
        candidates.append(gf.get("uuid"))
        if isinstance(gf.get("location"), dict):
            candidates.append((gf.get("location") or {}).get("uuid"))

    for candidate in candidates:
        value = _coerce_str(candidate)
        if value:
            return value
    return None

def _extract_bg_float(payload: dict, coords: dict, *names: str) -> float | None:
    for name in names:
        value = payload.get(name)
        if value is None:
            value = coords.get(name)
        if value is None and name == "lng":
            value = payload.get("lon") or coords.get("longitude")
        if value is None and name == "lon":
            value = payload.get("lng") or coords.get("longitude")
        if value is None and name == "lat":
            value = coords.get("latitude")
        if value is None and name in {"accuracy", "accuracy_m"}:
            value = coords.get("accuracy")
        if value is not None:
            try:
                return float(value)
            except (TypeError, ValueError):
                return None
    return None

def _resolve_bg_event_shift(payload: dict, device_uuid: str | None):
    employee_code = normalize_employee_code(
        payload.get("employee_code") or payload.get("username_code") or payload.get("employeeCode") or ""
    )
    expected_shift_id = payload.get("shift_id") or payload.get("shiftId")
    expected_store_id = payload.get("store_id") or payload.get("storeId")

    try:
        expected_shift_id = int(expected_shift_id) if expected_shift_id is not None else None
        expected_store_id = int(expected_store_id) if expected_store_id is not None else None
    except (TypeError, ValueError):
        return None, None, None, "invalid_expected_shift"

    emp = None
    if employee_code:
        emp = Employee.query.filter(func.lower(Employee.username_code) == employee_code).first()
        if emp and not emp.active:
            return emp, None, None, "inactive_employee"

    q = Shift.query.filter(Shift.clock_out.is_(None))
    if emp:
        q = q.filter(Shift.employee_id == emp.id)
    elif device_uuid:
        q = q.filter(Shift.clock_in_device_uuid == device_uuid)
    else:
        return None, None, None, "missing_employee_or_device"

    if expected_shift_id is not None:
        q = q.filter(Shift.id == expected_shift_id)

    open_shift = q.order_by(Shift.clock_in.desc()).first()
    if not open_shift:
        return emp, None, None, "no_open_shift"

    if not emp:
        emp = Employee.query.get(open_shift.employee_id)
        if emp and not emp.active:
            return emp, None, None, "inactive_employee"

    if expected_store_id is not None and open_shift.store_id != expected_store_id:
        return emp, open_shift, None, "open_shift_store_changed"

    store = Store.query.get(open_shift.store_id)
    if not store:
        return emp, open_shift, None, "store_not_found"

    return emp, open_shift, store, None

def _coerce_str(val, max_len: int = 120) -> str | None:
    if val is None:
        return None
    try:
        s = str(val).strip()
    except Exception:
        return None
    if not s:
        return None
    return s[:max_len]

def _json_bool_field(data: dict, name: str) -> tuple[bool | None, str | None]:
    if name not in data:
        return False, None
    value = data.get(name)
    if isinstance(value, bool):
        return value, None
    return None, f"{name}_must_be_boolean"

def _touch_employee_device(emp: "Employee", device_uuid: str | None, device_label: str | None):
    """
    Option C behavior: if device_uuid provided, overwrite employee.device_uuid.
    Never blocks clock-ins.
    """
    if not device_uuid:
        return
    try:
        emp.device_uuid = device_uuid
        if device_label:
            emp.device_label = device_label
        emp.device_last_seen_at = now_utc()
    except Exception:
        pass

def _auto_exit_clearance_required(accuracy_m: float | None = None) -> float:
    accuracy_part = 0.0
    if accuracy_m is not None:
        try:
            accuracy_part = max(0.0, min(float(accuracy_m), AUTO_EXIT_ACCURACY_MAX_M))
        except (TypeError, ValueError):
            accuracy_part = 0.0
    return AUTO_EXIT_OUTSIDE_BUFFER_M + accuracy_part

def _is_reliable_outside(store: "Store", dist_m: float, accuracy_m: float | None = None) -> bool:
    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        return False
    return dist_m > (store.geofence_radius_m + _auto_exit_clearance_required(accuracy_m))

def _cancel_pending_auto_exit(shift_id: int, reason: str):
    pending_records = (
        PendingAutoExit.query
        .filter(PendingAutoExit.shift_id == shift_id, PendingAutoExit.status.in_(["active", "candidate"]))
        .order_by(PendingAutoExit.created_at.desc())
        .all()
    )
    if not pending_records:
        return False
    now = now_utc()
    for pending in pending_records:
        pending.status = "cancelled"
        pending.cancel_reason = reason[:120]
        pending.updated_at = now
    db.session.commit()
    log_event("AUTO_EXIT_PENDING_CANCELLED", shift_id=shift_id, reason=reason, count=len(pending_records))
    return True

def _record_auto_exit_observation(
    emp: "Employee",
    open_shift: "Shift",
    store: "Store",
    lat: float,
    lng: float,
    accuracy_m: float | None,
    device_uuid: str | None,
    source: str,
):
    source = (source or "outside_location").strip().lower()
    dist_m = haversine_m(lat, lng, store.latitude, store.longitude)
    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        return None, "accuracy_too_low", dist_m

    if not _is_reliable_outside(store, dist_m, accuracy_m):
        _cancel_pending_auto_exit(open_shift.id, "inside_or_near_store")
        return None, "inside_or_near_store", dist_m

    now = now_utc()
    pending = (
        PendingAutoExit.query
        .filter(PendingAutoExit.shift_id == open_shift.id, PendingAutoExit.status.in_(["active", "candidate"]))
        .with_for_update()
        .order_by(PendingAutoExit.created_at.desc())
        .first()
    )

    starts_immediately = source == "geofence_exit"

    if pending:
        if pending.employee_id != emp.id or pending.store_id != store.id:
            return pending, "pending_shift_identity_mismatch", dist_m
        if pending.device_uuid and device_uuid and pending.device_uuid != device_uuid:
            return pending, "pending_device_mismatch", dist_m
        pending.last_seen_at = now
        pending.last_lat = lat
        pending.last_lng = lng
        pending.last_accuracy_m = accuracy_m
        pending.last_dist_m = float(dist_m)
        pending.outside_count = (pending.outside_count or 0) + 1
        pending.device_uuid = device_uuid or pending.device_uuid
        pending.source = source or pending.source
        pending.updated_at = now
        if pending.status == "candidate":
            age_seconds = (now - (pending.exit_at or pending.created_at or now)).total_seconds()
            if starts_immediately or age_seconds >= AUTO_EXIT_OUTSIDE_CONFIRM_SECONDS:
                pending.status = "active"
                pending.deadline_at = now + timedelta(seconds=AUTO_EXIT_GRACE_SECONDS)
    else:
        pending = PendingAutoExit(
            employee_id=emp.id,
            shift_id=open_shift.id,
            store_id=store.id,
            device_uuid=device_uuid,
            status="active" if starts_immediately else "candidate",
            source=source or "outside_location",
            exit_at=now,
            deadline_at=now + timedelta(seconds=AUTO_EXIT_GRACE_SECONDS),
            last_seen_at=now,
            last_lat=lat,
            last_lng=lng,
            last_accuracy_m=accuracy_m,
            last_dist_m=float(dist_m),
            outside_count=1,
        )
        db.session.add(pending)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        pending = (
            PendingAutoExit.query
            .filter(PendingAutoExit.shift_id == open_shift.id, PendingAutoExit.status.in_(["active", "candidate"]))
            .with_for_update()
            .order_by(PendingAutoExit.created_at.desc())
            .first()
        )
        if not pending:
            raise
        pending.last_seen_at = now
        pending.last_lat = lat
        pending.last_lng = lng
        pending.last_accuracy_m = accuracy_m
        pending.last_dist_m = float(dist_m)
        pending.outside_count = (pending.outside_count or 0) + 1
        pending.device_uuid = device_uuid or pending.device_uuid
        pending.updated_at = now
        if pending.status == "candidate":
            age_seconds = (now - (pending.exit_at or pending.created_at or now)).total_seconds()
            if starts_immediately or age_seconds >= AUTO_EXIT_OUTSIDE_CONFIRM_SECONDS:
                pending.status = "active"
                pending.deadline_at = now + timedelta(seconds=AUTO_EXIT_GRACE_SECONDS)
        db.session.commit()
    log_event(
        "AUTO_EXIT_PENDING_UPSERT",
        employee_id=emp.id,
        shift_id=open_shift.id,
        store_id=store.id,
        source=source,
        status=pending.status,
        outside_count=pending.outside_count,
        dist_m=round(dist_m, 1),
        accuracy_m=accuracy_m if accuracy_m is not None else "",
        deadline_at=pending.deadline_at.isoformat(),
    )
    if pending.status != "active":
        return pending, "outside_candidate_waiting_for_confirmation", dist_m
    return pending, None, dist_m

def _close_shift_auto_exit_from_pending(pending: "PendingAutoExit"):
    pending = (
        PendingAutoExit.query
        .filter(PendingAutoExit.id == pending.id)
        .with_for_update()
        .first()
    )
    if not pending or pending.status != "active" or not pending.deadline_at or pending.deadline_at > now_utc():
        return False, "pending_not_expired"

    open_shift = (
        Shift.query
        .filter(Shift.id == pending.shift_id, Shift.clock_out.is_(None))
        .with_for_update()
        .first()
    )
    if not open_shift or open_shift.clock_out is not None:
        pending.status = "cancelled"
        pending.cancel_reason = "shift_already_closed"
        pending.updated_at = now_utc()
        db.session.commit()
        return False, "shift_already_closed"

    store = Store.query.get(open_shift.store_id)
    emp = Employee.query.get(open_shift.employee_id)
    if not store or not emp or open_shift.store_id != pending.store_id:
        pending.status = "cancelled"
        pending.cancel_reason = "shift_store_or_employee_missing"
        pending.updated_at = now_utc()
        db.session.commit()
        return False, "shift_store_or_employee_missing"

    if pending.last_lat is None or pending.last_lng is None:
        return False, "missing_location"

    accuracy_m = pending.last_accuracy_m
    dist_m = haversine_m(pending.last_lat, pending.last_lng, store.latitude, store.longitude)
    if not _is_reliable_outside(store, dist_m, accuracy_m):
        pending.status = "cancelled"
        pending.cancel_reason = "inside_or_near_store_at_deadline"
        pending.updated_at = now_utc()
        db.session.commit()
        return False, "inside_or_near_store_at_deadline"

    old_in = open_shift.clock_in
    old_out = open_shift.clock_out
    now = now_utc()
    close_time = pending.deadline_at

    open_shift.clock_out = close_time
    open_shift.clock_out_lat = pending.last_lat
    open_shift.clock_out_lng = pending.last_lng
    open_shift.clock_out_device_uuid = pending.device_uuid
    open_shift.clock_out_source = "auto_exit"
    open_shift.closed_by_admin = False
    open_shift.admin_closed_by = "AUTO_EXIT"
    open_shift.admin_closed_at = now
    open_shift.admin_close_reason = (
        f"Auto Exit after {AUTO_EXIT_GRACE_SECONDS // 60} minute grace; "
        f"clock_out_at=deadline processed_at={now.isoformat()}; "
        f"dist={round(dist_m, 1)}m radius={store.geofence_radius_m}m "
        f"accuracy={round(accuracy_m, 1) if accuracy_m is not None else 'unknown'}m"
    )

    audit = ShiftEditAudit(
        shift_id=open_shift.id,
        action="auto_exit_close",
        editor="AUTO_EXIT",
        reason=open_shift.admin_close_reason,
        old_clock_in=old_in,
        old_clock_out=old_out,
        new_clock_in=open_shift.clock_in,
        new_clock_out=open_shift.clock_out,
    )
    pending.status = "closed"
    pending.closed_at = now
    pending.updated_at = now
    db.session.add(audit)
    db.session.commit()
    log_event(
        "AUTO_EXIT_PENDING_CLOSED",
        employee_id=emp.id,
        shift_id=open_shift.id,
        store_id=store.id,
        dist_m=round(dist_m, 1),
        accuracy_m=accuracy_m if accuracy_m is not None else "",
    )
    return True, "closed"

def process_expired_pending_auto_exits(limit: int = 50):
    now = now_utc()
    expired = (
        PendingAutoExit.query
        .filter(PendingAutoExit.status == "active", PendingAutoExit.deadline_at <= now)
        .with_for_update(skip_locked=True)
        .order_by(PendingAutoExit.deadline_at.asc())
        .limit(limit)
        .all()
    )
    results = []
    for pending in expired:
        try:
            results.append((pending.id, *_close_shift_auto_exit_from_pending(pending)))
        except Exception as exc:
            db.session.rollback()
            app.logger.exception("AUTO_EXIT_PENDING_PROCESS_FAILED id=%s error=%s", pending.id, exc)
            results.append((pending.id, False, "error"))
    return results

def process_expired_pending_auto_exits_best_effort(context: str = ""):
    try:
        return process_expired_pending_auto_exits()
    except Exception:
        app.logger.exception("AUTO_EXIT_PENDING_OPPORTUNISTIC_PROCESS_FAILED context=%s", context)
        return []

_auto_exit_worker_started = False

def start_auto_exit_worker():
    # Best-effort only: Render/Gunicorn may run multiple processes or restart.
    # Correctness comes from database-backed pending exits plus opportunistic
    # processing and the protected maintenance endpoint for external scheduling.
    global _auto_exit_worker_started
    if _auto_exit_worker_started:
        return
    if (os.environ.get("AUTO_EXIT_WORKER_ENABLED") or "0").strip().lower() not in {"1", "true", "yes"}:
        return
    _auto_exit_worker_started = True

    def _worker():
        while True:
            time.sleep(AUTO_EXIT_WORKER_INTERVAL_SECONDS)
            try:
                with app.app_context():
                    process_expired_pending_auto_exits()
            except Exception:
                app.logger.exception("AUTO_EXIT_WORKER_ERROR")

    t = threading.Thread(target=_worker, name="clockin-auto-exit-worker", daemon=True)
    t.start()
    app.logger.info("Auto-exit worker started interval=%ss", AUTO_EXIT_WORKER_INTERVAL_SECONDS)

def _device_has_other_open_shift(device_uuid: str, employee_id: int) -> "Shift | None":
    """
    Prevent the obvious abuse: one phone can't have an open shift for Employee A
    while Employee B tries to clock in on same device.
    """
    if not device_uuid:
        return None
    return (
        Shift.query
        .filter(
            Shift.clock_out.is_(None),
            Shift.clock_in_device_uuid == device_uuid,
            Shift.employee_id != employee_id
        )
        .order_by(Shift.clock_in.desc())
        .first()
    )

# Make helpers available in templates
@app.context_processor
def inject_helpers():
    return dict(
        fmt_dt=fmt_dt,
        shift_minutes=shift_minutes,
        minutes_to_human=minutes_to_human,
        minutes_to_short=minutes_to_short,
        minutes_to_decimal_hours=minutes_to_decimal_hours,
        shift_clock_out_source=shift_clock_out_source,
        shift_clock_out_label=shift_clock_out_label,
        shift_clock_out_distance_m=shift_clock_out_distance_m,
        shift_clock_out_outside_geofence=shift_clock_out_outside_geofence,
        active_pending_auto_exit_for_shift=active_pending_auto_exit_for_shift,
        pending_auto_exit_minutes_remaining=pending_auto_exit_minutes_remaining,
        suggest_employee_username=suggest_employee_username
    )

# -----------------------------
# ✅ Option B-safe: add missing columns without migrations
# -----------------------------
def _ensure_column(table_name: str, column_name: str, sql_type: str):
    """
    Best-effort: add a column if it doesn't exist.
    Works for SQLite and Postgres for simple ADD COLUMN cases.
    Race-safe on Render: ignores "already exists" errors.
    """
    try:
        bind = db.engine
        dialect = bind.dialect.name

        exists = False

        if dialect == "postgresql":
            q = text("""
                SELECT 1
                FROM information_schema.columns
                WHERE table_name = :t AND column_name = :c
                LIMIT 1
            """)
            row = db.session.execute(q, {"t": table_name, "c": column_name}).first()
            exists = bool(row)

        elif dialect == "sqlite":
            q = text(f"PRAGMA table_info({table_name})")
            rows = db.session.execute(q).fetchall()
            exists = any((r[1] == column_name) for r in rows)  # r[1] = name

        if exists:
            return

        db.session.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {sql_type}"))
        db.session.commit()
        app.logger.info("Added missing column %s.%s", table_name, column_name)

    except Exception as e:
        db.session.rollback()
        msg = str(e).lower()
        if "already exists" in msg or "duplicate" in msg:
            app.logger.info("Column already exists (race): %s.%s", table_name, column_name)
            return
        app.logger.exception("Could not ensure column %s.%s", table_name, column_name)

def _ensure_unique_employee_code_index():
    try:
        dialect = db.engine.dialect.name
        if dialect == "sqlite":
            sql = """
            CREATE UNIQUE INDEX IF NOT EXISTS ix_employees_username_code
            ON employees (username_code)
            WHERE username_code IS NOT NULL AND username_code != ''
            """
        elif dialect == "postgresql":
            sql = """
            CREATE UNIQUE INDEX IF NOT EXISTS ix_employees_username_code
            ON employees (username_code)
            WHERE username_code IS NOT NULL AND username_code <> ''
            """
        else:
            sql = "CREATE UNIQUE INDEX ix_employees_username_code ON employees (username_code)"
        db.session.execute(text(sql))
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception("Could not ensure unique employee username/code index")

def _ensure_pending_auto_exit_indexes():
    try:
        statements = [
            """
            CREATE INDEX IF NOT EXISTS ix_pending_auto_exits_shift_status
            ON pending_auto_exits (shift_id, status)
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_pending_auto_exits_status_deadline
            ON pending_auto_exits (status, deadline_at)
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_pending_auto_exits_employee_status
            ON pending_auto_exits (employee_id, status)
            """,
            """
            CREATE INDEX IF NOT EXISTS ix_pending_auto_exits_store_status
            ON pending_auto_exits (store_id, status)
            """,
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_pending_auto_exits_open_shift
            ON pending_auto_exits (shift_id)
            WHERE status IN ('active', 'candidate')
            """,
        ]
        for sql in statements:
            db.session.execute(text(sql))
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception("Could not ensure pending auto-exit indexes")

# -----------------------------
# Create tables on startup (Option B)
# -----------------------------
with app.app_context():
    try:
        db.create_all()
        app.logger.info("DB create_all OK")
    except Exception as e:
        app.logger.exception("DB create_all failed: %s", e)

    # Ensure new Option 2 columns exist (no migrations needed)
    _ensure_column("employees", "username_code", "VARCHAR(80)")
    _ensure_column("employees", "device_uuid", "VARCHAR(120)")
    _ensure_column("employees", "device_label", "VARCHAR(120)")
    _ensure_column("employees", "device_last_seen_at", "TIMESTAMP")
    _ensure_unique_employee_code_index()

    _ensure_column("shifts", "clock_in_device_uuid", "VARCHAR(120)")
    _ensure_column("shifts", "clock_out_device_uuid", "VARCHAR(120)")
    _ensure_column("shifts", "clock_out_source", "VARCHAR(32)")
    _ensure_pending_auto_exit_indexes()
    start_auto_exit_worker()

# -----------------------------
# Fingerprint (DEBUG)
# -----------------------------
@app.get("/__fingerprint__")
def fingerprint():
    return "clockin_app LIVE fingerprint 2026-02-16"

# -----------------------------
# Optional: favicon
# -----------------------------
@app.get("/favicon.ico")
def favicon():
    return ("", 204)

# -----------------------------
# DEV endpoints (locked down)
# -----------------------------
@app.get("/dev/db-info")
def dev_db_info():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    uri = app.config.get("SQLALCHEMY_DATABASE_URI", "")
    return jsonify({
        "ok": True,
        "db_uri": uri,
        "store_count": Store.query.count(),
    })

@app.get("/dev/routes")
def dev_routes():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code
    return jsonify(sorted([str(r) for r in app.url_map.iter_rules()]))

@app.get("/dev/export-stores", endpoint="dev_export_stores_v2")
def dev_export_stores():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    stores = Store.query.order_by(Store.id.asc()).all()
    return jsonify({
        "ok": True,
        "stores": [
            {
                "name": s.name,
                "qr_token": s.qr_token,
                "latitude": s.latitude,
                "longitude": s.longitude,
                "geofence_radius_m": s.geofence_radius_m,
            }
            for s in stores
        ]
    })

@app.get("/dev/export-employees", endpoint="dev_export_employees_v2")
def dev_export_employees():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    emps = Employee.query.order_by(Employee.id.asc()).all()
    return jsonify({
        "ok": True,
        "employees": [
            {
                "name": e.name,
                "username_code": e.username_code or "",
                "suggested_username_code": suggest_employee_username(e.name),
                "pin": e.pin,
                "active": bool(e.active)
            }
            for e in emps
        ]
    })

@app.post("/dev/import-stores")
def dev_import_stores():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}
    stores = data.get("stores") or []
    if not isinstance(stores, list):
        return jsonify({"ok": False, "error": "stores_must_be_list"}), 400

    upserted = 0
    for s in stores:
        name = (s.get("name") or "").strip()
        qr_token = normalize_store_code(s.get("qr_token") or "")
        lat = s.get("latitude")
        lon = s.get("longitude")
        radius = s.get("geofence_radius_m", 150)

        if not name or not qr_token or lat is None or lon is None:
            continue

        try:
            lat = float(lat)
            lon = float(lon)
            radius = int(radius)
        except (TypeError, ValueError):
            continue

        existing = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
        if existing:
            existing.name = name
            existing.latitude = lat
            existing.longitude = lon
            existing.geofence_radius_m = radius
        else:
            db.session.add(Store(
                name=name,
                qr_token=qr_token,
                latitude=lat,
                longitude=lon,
                geofence_radius_m=radius,
                created_at=now_utc()
            ))
        upserted += 1

    db.session.commit()
    return jsonify({"ok": True, "imported_or_updated": upserted})

@app.post("/dev/import-employees")
def dev_import_employees():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}
    employees = data.get("employees") or []
    if not isinstance(employees, list):
        return jsonify({"ok": False, "error": "employees_must_be_list"}), 400

    upserted = 0
    for e in employees:
        name = (e.get("name") or "").strip()
        username_code = normalize_employee_code(e.get("username_code") or e.get("employee_code") or "")
        pin = (e.get("pin") or "").strip()
        active = bool(e.get("active", True))

        if not name or not pin:
            continue

        if not username_code:
            username_code = unique_employee_code_from_name(name)

        existing = Employee.query.filter_by(pin=pin).first()
        if existing:
            existing.name = name
            existing.username_code = username_code
            existing.active = active
        else:
            db.session.add(Employee(
                name=name,
                username_code=username_code,
                pin=pin,
                active=active,
                created_at=now_utc()
            ))
        upserted += 1

    db.session.commit()
    return jsonify({"ok": True, "imported_or_updated": upserted})

@app.post("/dev/add-store")
def dev_add_store():
    ok, err = _dev_guard()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}

    name = (data.get("name") or "").strip()
    qr_token = normalize_store_code(data.get("qr_token") or "")
    lat = data.get("lat")
    lon = data.get("lon")
    radius = data.get("geofence_radius_m", 200)

    if not name or not qr_token or lat is None or lon is None:
        return jsonify({"ok": False, "error": "missing_fields", "required": ["name", "qr_token", "lat", "lon"]}), 400

    try:
        lat = float(lat)
        lon = float(lon)
        radius = int(radius)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_values"}), 400

    store = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
    if store:
        store.name = name
        store.latitude = lat
        store.longitude = lon
        store.geofence_radius_m = radius
    else:
        store = Store(
            name=name,
            qr_token=qr_token,
            latitude=lat,
            longitude=lon,
            geofence_radius_m=radius
        )
        db.session.add(store)

    db.session.commit()
    return jsonify({"ok": True, "store_id": store.id, "name": store.name})

# -----------------------------
# Store Suggest API (Autocomplete)
# -----------------------------
@app.get("/api/stores/suggest")
def api_stores_suggest():
    """
    Autocomplete support for employee store-code entry.
    Query: /api/stores/suggest?q=rea
    Returns: [{code, name}] (code is qr_token)
    """
    q = (request.args.get("q") or "").strip()
    if not q:
        return jsonify([])

    ql = q.lower()
    matches = (
        Store.query
        .filter(
            (func.lower(Store.qr_token).like(f"%{ql}%")) |
            (func.lower(Store.name).like(f"%{ql}%"))
        )
        .order_by(Store.name.asc())
        .limit(10)
        .all()
    )

    return jsonify([{"code": s.qr_token, "name": s.name} for s in matches])

@app.get("/api/stores/all")
def api_stores_all():
    """
    Returns all stores for the mobile store picker.
    Public (no auth). Only exposes store name + code.
    """
    stores = Store.query.order_by(Store.name.asc()).all()
    return jsonify([{"code": s.qr_token, "name": s.name} for s in stores])

# -----------------------------
# ✅ Mobile identity + geofence endpoints (Option 2)
# -----------------------------
@app.post("/api/mobile/me")
def api_mobile_me():
    """
    Body: { username_code, pin, device_uuid?, device_label? }
    """
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code
    process_expired_pending_auto_exits_best_effort("mobile_me")

    data = request.get_json(silent=True) or {}
    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin:
        return jsonify({"ok": False, "error": "missing_pin"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    _touch_employee_device(emp, device_uuid, device_label)
    db.session.commit()

    return jsonify({
        "ok": True,
        "employee": employee_payload(emp),
        "server_time_utc": now_utc().isoformat() + "Z"
    })

@app.post("/api/mobile/status")
def api_mobile_status():
    """
    Returns employee identity + current open shift (if any).
    Body: { username_code, pin, device_uuid?, device_label? }
    """
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code
    process_expired_pending_auto_exits_best_effort("mobile_status")

    data = request.get_json(silent=True) or {}
    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin:
        return jsonify({"ok": False, "error": "missing_pin"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    _touch_employee_device(emp, device_uuid, device_label)

    open_shift = (
        Shift.query
        .filter(Shift.employee_id == emp.id, Shift.clock_out.is_(None))
        .order_by(Shift.clock_in.desc())
        .first()
    )

    payload = {
        "ok": True,
        "employee": employee_payload(emp),
        "server_time_utc": now_utc().isoformat() + "Z",
        "open_shift": None,
    }

    if open_shift:
        store = Store.query.get(open_shift.store_id)
        payload["open_shift"] = {
            "shift_id": open_shift.id,
            "store_id": open_shift.store_id,
            "store_name": store.name if store else "",
            "clock_in_utc": open_shift.clock_in.isoformat() + "Z",
            "clock_in_local": fmt_dt(open_shift.clock_in),
            "closed_by_admin": bool(open_shift.closed_by_admin),
        }

    db.session.commit()
    return jsonify(payload), 200

@app.post("/api/mobile/clock-in")
def api_mobile_clock_in():
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code
    process_expired_pending_auto_exits_best_effort("mobile_clock_in")

    data = request.get_json(silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    qr_token = normalize_store_code(
        data.get("qr_token") or data.get("store_code") or ""
    )

    lat = data.get("lat")
    lon = data.get("lon")
    accuracy_m = data.get("accuracy_m")
    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin or not qr_token or lat is None or lon is None:
        return jsonify({
            "ok": False,
            "error": "missing_required_fields",
            "required": ["pin", "qr_token", "lat", "lon"]
        }), 400

    try:
        lat = float(lat)
        lon = float(lon)
        if accuracy_m is not None:
            accuracy_m = float(accuracy_m)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_location"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    selected_store = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
    if not selected_store:
        return jsonify({"ok": False, "error": "invalid_store_code"}), 404

    existing = Shift.query.filter(
        Shift.employee_id == emp.id,
        Shift.clock_out.is_(None)
    ).first()

    if existing:
        return jsonify({"ok": False, "error": "already_clocked_in"}), 409

    if device_uuid:
        other = _device_has_other_open_shift(device_uuid, emp.id)
        if other:
            return jsonify({"ok": False, "error": "device_in_use"}), 409

    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        return jsonify({
            "ok": False,
            "error": "accuracy_too_low",
            "message": "GPS accuracy is too low. Step outside and try again.",
            "accuracy_m": accuracy_m
        }), 403

    dist_m = haversine_m(
        lat,
        lon,
        selected_store.latitude,
        selected_store.longitude
    )

    if dist_m > selected_store.geofence_radius_m:
        return jsonify({
            "ok": False,
            "error": "outside_selected_store_geofence",
            "message": "You are not at the selected store location.",
            "store_name": selected_store.name,
            "distance_m": round(dist_m, 1),
            "required_radius_m": selected_store.geofence_radius_m
        }), 403

    _touch_employee_device(emp, device_uuid, device_label)

    shift = Shift(
        employee_id=emp.id,
        store_id=selected_store.id,
        clock_in=now_utc(),
        clock_in_lat=lat,
        clock_in_lng=lon,
        clock_in_device_uuid=device_uuid
    )

    db.session.add(shift)
    db.session.commit()

    return jsonify({
        "ok": True,
        "shift_id": shift.id,
        "employee_id": emp.id,
        "employee_name": emp.name,
        "store_id": selected_store.id,
        "store_name": selected_store.name,
        "distance_m": round(dist_m, 1),
        "clock_in_utc": shift.clock_in.isoformat() + "Z"
    }), 200

@app.post("/api/mobile/clock-out")
def api_mobile_clock_out():
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code
    process_expired_pending_auto_exits_best_effort("mobile_clock_out")

    data = request.get_json(silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    lat = data.get("lat")
    lon = data.get("lon")
    accuracy_m = data.get("accuracy_m")
    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))
    allow_outside_geofence, bool_error = _json_bool_field(data, "allow_outside_geofence")
    if bool_error:
        return jsonify({"ok": False, "error": bool_error}), 400

    if not pin or lat is None or lon is None:
        return jsonify({"ok": False, "error": "missing_required_fields"}), 400

    try:
        lat = float(lat)
        lon = float(lon)
        if accuracy_m is not None:
            accuracy_m = float(accuracy_m)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_location"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    open_shift = (
        Shift.query
        .filter(Shift.employee_id == emp.id, Shift.clock_out.is_(None))
        .order_by(Shift.clock_in.desc())
        .first()
    )

    if not open_shift:
        return jsonify({"ok": False, "error": "no_open_shift"}), 409

    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        return jsonify({
            "ok": False,
            "error": "accuracy_too_low",
            "message": "GPS accuracy is too low. Step outside and try again.",
            "accuracy_m": accuracy_m,
        }), 403

    store = Store.query.get(open_shift.store_id)
    if not store:
        return jsonify({"ok": False, "error": "store_not_found"}), 500

    dist_m = haversine_m(lat, lon, store.latitude, store.longitude)
    outside_geofence = dist_m > store.geofence_radius_m
    if outside_geofence and not allow_outside_geofence:
        return jsonify({
            "ok": False,
            "error": "outside_store_geofence",
            "requires_confirmation": True,
            "message": "You appear to be outside the store location. Clock out anyway?",
            "dist_m": round(dist_m, 1),
            "radius_m": float(store.geofence_radius_m),
            "accuracy_m": accuracy_m,
        }), 409

    _touch_employee_device(emp, device_uuid, device_label)

    old_in = open_shift.clock_in
    old_out = open_shift.clock_out

    open_shift.clock_out = now_utc()
    open_shift.clock_out_lat = lat
    open_shift.clock_out_lng = lon
    open_shift.clock_out_device_uuid = device_uuid
    open_shift.clock_out_source = "employee"

    if outside_geofence:
        audit = ShiftEditAudit(
            shift_id=open_shift.id,
            action="employee_clock_out_outside",
            editor="MOBILE",
            reason=(
                f"outside_geofence=true confirmation=true source=mobile "
                f"lat={lat} lon={lon} dist_m={round(dist_m, 1)} "
                f"radius_m={store.geofence_radius_m} "
                f"accuracy_m={round(accuracy_m, 1) if accuracy_m is not None else 'unknown'}"
            ),
            old_clock_in=old_in,
            old_clock_out=old_out,
            new_clock_in=open_shift.clock_in,
            new_clock_out=open_shift.clock_out,
        )
        db.session.add(audit)

    _cancel_pending_auto_exit(open_shift.id, "manual_clock_out")

    db.session.commit()

    minutes = shift_minutes(open_shift)

    return jsonify({
        "ok": True,
        "shift_id": open_shift.id,
        "clock_out_utc": open_shift.clock_out.isoformat() + "Z",
        "outside_geofence": outside_geofence,
        "dist_m": round(dist_m, 1),
        "radius_m": float(store.geofence_radius_m),
        "minutes": minutes,
        "human": minutes_to_human(minutes)
    })

@app.post("/api/mobile/auto-exit-close")
def api_mobile_auto_exit_close():
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    lat = data.get("lat")
    lon = data.get("lon")
    accuracy_m = data.get("accuracy_m")
    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))
    expected_shift_id = data.get("shift_id")
    expected_store_id = data.get("store_id")

    # optional: reason from app
    reason = (data.get("reason") or "Auto-close after EXIT").strip()

    if not pin or lat is None or lon is None:
        return jsonify({"ok": False, "error": "missing_required_fields"}), 400

    try:
        lat = float(lat)
        lon = float(lon)
        if accuracy_m is not None:
            accuracy_m = float(accuracy_m)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_location"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    # Open shift required
    open_shift = (
        Shift.query
        .filter(Shift.employee_id == emp.id, Shift.clock_out.is_(None))
        .order_by(Shift.clock_in.desc())
        .first()
    )
    if not open_shift:
        return jsonify({"ok": True, "already_closed": True, "message": "No open shift."}), 200

    try:
        expected_shift_id = int(expected_shift_id) if expected_shift_id is not None else None
        expected_store_id = int(expected_store_id) if expected_store_id is not None else None
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_expected_shift"}), 400

    if expected_shift_id is not None and open_shift.id != expected_shift_id:
        return jsonify({
            "ok": False,
            "error": "open_shift_changed",
            "message": "Open shift changed before auto-close.",
            "open_shift_id": open_shift.id,
        }), 409

    if expected_store_id is not None and open_shift.store_id != expected_store_id:
        return jsonify({
            "ok": False,
            "error": "open_shift_store_changed",
            "message": "Open shift store changed before auto-close.",
            "open_shift_id": open_shift.id,
        }), 409

    store = Store.query.get(open_shift.store_id)
    if not store:
        return jsonify({"ok": False, "error": "store_not_found"}), 500

    # Distance check
    dist_m = haversine_m(lat, lon, store.latitude, store.longitude)

    # Accuracy gate (prevent bad GPS closing someone incorrectly)
    # Match your validate-location gate style
    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        return jsonify({
            "ok": False,
            "error": "accuracy_too_low",
            "message": "GPS accuracy too low to auto-close. Try again.",
            "accuracy_m": accuracy_m
        }), 409

    # Only allow auto-close if OUTSIDE radius (with a little buffer)
    buffer_m = AUTO_EXIT_OUTSIDE_BUFFER_M
    if not _is_reliable_outside(store, dist_m, accuracy_m):
        return jsonify({
            "ok": False,
            "error": "still_inside_or_near_store",
            "dist_m": float(dist_m),
            "radius_m": float(store.geofence_radius_m),
            "buffer_m": buffer_m
        }), 409

    _touch_employee_device(emp, device_uuid, device_label)

    pending, pending_error, _ = _record_auto_exit_observation(
        emp,
        open_shift,
        store,
        lat,
        lon,
        accuracy_m,
        device_uuid,
        "legacy_auto_exit_close",
    )
    results = process_expired_pending_auto_exits()
    closed = any(ok and result == "closed" for _, ok, result in results)

    log_event(
        "AUTO_EXIT_LEGACY_OBSERVATION",
        employee_id=emp.id,
        shift_id=open_shift.id,
        store_id=store.id,
        dist_m=round(dist_m, 1),
        radius_m=store.geofence_radius_m,
        accuracy_m=accuracy_m if accuracy_m is not None else "",
        device_uuid=device_uuid or "",
        pending_status=getattr(pending, "status", "") if pending else "",
        pending_error=pending_error or "",
        closed=closed,
    )

    return jsonify({
        "ok": True,
        "shift_id": open_shift.id,
        "store_name": store.name,
        "dist_m": round(dist_m, 1),
        "pending_status": getattr(pending, "status", None),
        "closed": closed,
        "message": "Auto-exit observation received."
    }), 200

@app.post("/api/mobile/geofences")
def api_mobile_geofences():
    """
    Returns the *real* store geofence based on store_code.
    Body: { username_code, pin, qr_token, device_uuid? }  (pin is required; device_uuid stored for last-seen)
    """
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    qr_token = normalize_store_code((data.get("qr_token") or "").strip())

    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin or not qr_token:
        return jsonify({"ok": False, "error": "missing_pin_or_store_code"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    store = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
    if not store:
        return jsonify({"ok": False, "error": "invalid_store_code"}), 404

    _touch_employee_device(emp, device_uuid, device_label)
    db.session.commit()

    geofences = [{
        "identifier": f"store_{store.id}",
        "latitude": float(store.latitude),
        "longitude": float(store.longitude),
        "radius": int(store.geofence_radius_m),
        "notifyOnEntry": True,
        "notifyOnExit": True
    }]

    return jsonify({
        "ok": True,
        "store": {"id": store.id, "name": store.name, "code": store.qr_token, "radius_m": store.geofence_radius_m},
        "geofences": geofences
    })

# -----------------------------
# ✅ Mobile event ingest (Transistorsoft BG Geolocation)
# -----------------------------
@app.post("/api/mobile/bg/event")
def api_mobile_bg_event():
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({"ok": False, "error": "invalid_json"}), 400

    event_type = (payload.get("event") or payload.get("name") or payload.get("type") or "unknown")
    event_type = str(event_type).strip().lower() or "unknown"

    loc, coords = _extract_location_coords(payload)

    device_uuid = _extract_device_uuid_from_bg_payload(payload, loc)

    is_moving = payload.get("is_moving")
    if is_moving is None and isinstance(loc, dict):
        is_moving = loc.get("is_moving")

    lat = _extract_bg_float(payload, coords, "lat", "latitude")
    lng = _extract_bg_float(payload, coords, "lng", "lon", "longitude")
    accuracy = _extract_bg_float(payload, coords, "accuracy_m", "accuracy")

    event_at = _extract_event_at(payload, loc)
    ping_created = False
    ping_id = None
    ping_skipped = None

    try:
        evt = MobileEvent(
            event_type=event_type,
            device_uuid=device_uuid,
            is_moving=bool(is_moving) if isinstance(is_moving, bool) else None,
            lat=lat,
            lng=lng,
            accuracy=accuracy,
            event_at=event_at,
            received_at=now_utc(),
            raw_json=_safe_json_dumps(payload),
        )
        db.session.add(evt)
        db.session.commit()
    except Exception:
        app.logger.exception("MOBILE_BG_EVENT_SAVE_FAILED")
        return jsonify({"ok": False, "error": "db_error"}), 500

    try:
        emp, open_shift, store, resolve_error = _resolve_bg_event_shift(payload, device_uuid)
        if resolve_error:
            ping_skipped = resolve_error
        elif lat is None or lng is None:
            ping_skipped = "missing_location"
        elif not emp or not open_shift or not store:
            ping_skipped = "missing_open_shift"
        else:
            geofence_action = ""
            if isinstance(payload.get("geofence"), dict):
                geofence_action = str(
                    payload.get("geofence_action")
                    or (payload.get("geofence") or {}).get("action")
                    or ""
                ).strip().upper()
            elif payload.get("geofence_action"):
                geofence_action = str(payload.get("geofence_action") or "").strip().upper()

            important_geofence = event_type == "geofence" and geofence_action in {"ENTER", "EXIT"}
            min_interval = BG_LOCATION_PING_INTERVAL
            dist_m = haversine_m(lat, lng, store.latitude, store.longitude)
            inside_radius = bool(dist_m <= store.geofence_radius_m)
            event_is_stale = bool(
                event_at and (now_utc() - event_at).total_seconds() > AUTO_EXIT_MAX_EVENT_AGE_SECONDS
            )

            if geofence_action == "ENTER" or inside_radius:
                _cancel_pending_auto_exit(open_shift.id, "geofence_enter" if geofence_action == "ENTER" else "inside_location")
            elif event_is_stale:
                log_event(
                    "AUTO_EXIT_PENDING_SKIPPED",
                    employee_id=emp.id,
                    shift_id=open_shift.id,
                    store_id=store.id,
                    reason="stale_event",
                    event_type=event_type,
                    geofence_action=geofence_action,
                    dist_m=round(dist_m, 1),
                    event_at=event_at.isoformat(),
                )
            elif important_geofence or event_type in {"location", "heartbeat", "motionchange"}:
                pending, pending_error, _ = _record_auto_exit_observation(
                    emp,
                    open_shift,
                    store,
                    lat,
                    lng,
                    accuracy,
                    device_uuid,
                    "geofence_exit" if geofence_action == "EXIT" else event_type,
                )
                if pending_error:
                    log_event(
                        "AUTO_EXIT_PENDING_SKIPPED",
                        employee_id=emp.id,
                        shift_id=open_shift.id,
                        store_id=store.id,
                        reason=pending_error,
                        event_type=event_type,
                        geofence_action=geofence_action,
                        dist_m=round(dist_m, 1),
                        accuracy_m=accuracy if accuracy is not None else "",
                    )

            cutoff = now_utc() - min_interval
            recent_ping = None
            if not important_geofence:
                recent_ping = (
                    LocationPing.query
                    .filter(LocationPing.shift_id == open_shift.id, LocationPing.created_at >= cutoff)
                    .order_by(LocationPing.created_at.desc())
                    .first()
                )

            if recent_ping:
                ping_skipped = "rate_limited"
            else:
                ping = LocationPing(
                    employee_id=emp.id,
                    shift_id=open_shift.id,
                    store_id=store.id,
                    lat=lat,
                    lng=lng,
                    dist_m=float(dist_m),
                    inside_radius=inside_radius,
                    created_at=now_utc(),
                )
                db.session.add(ping)
                db.session.commit()
                ping_created = True
                ping_id = ping.id
                log_event(
                    "BG_PING_OK",
                    employee_id=emp.id,
                    shift_id=open_shift.id,
                    store_id=store.id,
                    event_type=event_type,
                    dist_m=round(dist_m, 1),
                    inside=ping.inside_radius,
                    device_uuid=device_uuid or "",
                )

            process_expired_pending_auto_exits_best_effort("mobile_bg_event")
    except Exception:
        db.session.rollback()
        ping_skipped = "ping_processing_error"
        app.logger.exception("MOBILE_BG_EVENT_PING_PROCESSING_FAILED event_id=%s", evt.id)

    return jsonify({
        "ok": True,
        "id": evt.id,
        "ping_created": ping_created,
        "ping_id": ping_id,
        "ping_skipped": ping_skipped,
    })

@app.post("/api/maintenance/auto-exit/process")
def api_maintenance_auto_exit_process():
    expected = (os.environ.get("AUTO_EXIT_MAINTENANCE_TOKEN") or "").strip()
    provided = (request.headers.get("X-Auto-Exit-Token") or "").strip()
    auth = (request.headers.get("Authorization") or "").strip()
    if not provided and auth.lower().startswith("bearer "):
        provided = auth[7:].strip()

    if not expected:
        return jsonify({"ok": False, "error": "maintenance_not_configured"}), 503
    if not provided or not hmac.compare_digest(provided, expected):
        return jsonify({"ok": False, "error": "unauthorized"}), 401

    results = process_expired_pending_auto_exits()
    closed = sum(1 for _, ok, reason in results if ok and reason == "closed")
    skipped = len(results) - closed
    return jsonify({
        "ok": True,
        "processed": len(results),
        "closed": closed,
        "skipped": skipped,
    })

@app.post("/api/mobile/report-issue")
def api_mobile_report_issue():
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    if not pin:
        return jsonify({"ok": False, "error": "missing_pin"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    msg = (data.get("message") or "").strip() or None
    payload = data.get("payload") or {}
    if not isinstance(payload, dict):
        payload = {"_raw": payload}

    # Try to attach store_id / shift_id if possible
    store_id = None
    shift_id = None

    try:
        store_obj = payload.get("store") if isinstance(payload.get("store"), dict) else {}
        store_code = normalize_store_code(store_obj.get("code") or "")
        if store_code:
            s = Store.query.filter(func.lower(Store.qr_token) == store_code).first()
            if s:
                store_id = s.id
    except Exception:
        pass

    try:
        open_shift = (
            Shift.query
            .filter(Shift.employee_id == emp.id, Shift.clock_out.is_(None))
            .order_by(Shift.clock_in.desc())
            .first()
        )
        if open_shift:
            shift_id = open_shift.id
            if not store_id:
                store_id = open_shift.store_id
    except Exception:
        pass

    try:
        report = MobileIssueReport(
            employee_id=emp.id,
            store_id=store_id,
            shift_id=shift_id,
            message=msg,
            payload_json=_safe_json_dumps(payload),
            status="open",
            created_at=now_utc(),
        )
        db.session.add(report)
        db.session.commit()
    except Exception:
        db.session.rollback()
        app.logger.exception("MOBILE_ISSUE_SAVE_FAILED")
        return jsonify({"ok": False, "error": "db_error"}), 500

    # Optional log line
    app.logger.warning(f"[MOBILE ISSUE] id={report.id} emp={emp.id} {emp.name} store_id={store_id} shift_id={shift_id}")

    return jsonify({"ok": True, "id": report.id})

@app.post("/api/mobile/bg/locations")
def api_mobile_bg_locations_bulk():
    """
    Optional bulk endpoint if you configure BG to POST arrays of locations.
    """
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    payload = request.get_json(silent=True) or {}
    if not isinstance(payload, dict):
        return jsonify({"ok": False, "error": "invalid_json"}), 400

    locations = payload.get("locations")
    if not isinstance(locations, list):
        return jsonify({"ok": False, "error": "expected_locations_array"}), 400

    device_uuid = payload.get("uuid")
    if device_uuid is not None:
        device_uuid = str(device_uuid)

    saved = 0
    try:
        for item in locations:
            if not isinstance(item, dict):
                continue

            coords = item.get("coords") if isinstance(item.get("coords"), dict) else {}
            ts_ms = item.get("timestamp")
            event_at = None
            if isinstance(ts_ms, (int, float)) and ts_ms > 0:
                try:
                    event_at = datetime.utcfromtimestamp(ts_ms / 1000.0)
                except Exception:
                    event_at = None

            evt = MobileEvent(
                event_type="location",
                device_uuid=str(item.get("uuid") or device_uuid) if (item.get("uuid") or device_uuid) else None,
                is_moving=bool(item.get("is_moving")) if isinstance(item.get("is_moving"), bool) else None,
                lat=float(coords.get("latitude")) if isinstance(coords.get("latitude"), (int, float)) else None,
                lng=float(coords.get("longitude")) if isinstance(coords.get("longitude"), (int, float)) else None,
                accuracy=float(coords.get("accuracy")) if isinstance(coords.get("accuracy"), (int, float)) else None,
                event_at=event_at,
                received_at=now_utc(),
                raw_json=_safe_json_dumps(item),
            )
            db.session.add(evt)
            saved += 1

        db.session.commit()
    except Exception:
        app.logger.exception("MOBILE_BG_LOCATIONS_SAVE_FAILED")
        return jsonify({"ok": False, "error": "db_error"}), 500

    return jsonify({"ok": True, "saved": saved})

# -----------------------------
# Legacy-ish mobile endpoints (now token protected)
# -----------------------------
@app.post("/mobile/validate-location")
def mobile_validate_location():
    ok, err = _require_mobile_auth()
    if not ok:
        msg, code = err
        return jsonify({"ok": False, "error": msg}), code

    data = request.get_json(silent=True) or {}

    lat = data.get("lat")
    lon = data.get("lon")
    accuracy_m = data.get("accuracy_m")

    if lat is None or lon is None:
        return jsonify({"ok": False, "error": "missing_lat_lon"}), 400

    try:
        lat = float(lat)
        lon = float(lon)
        if accuracy_m is not None:
            accuracy_m = float(accuracy_m)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "invalid_lat_lon"}), 400

    result = find_store_for_location(lat, lon, accuracy_m)

    if not result.get("ok"):
        return jsonify(result), 200

    store = result["store"]
    return jsonify({
        "ok": True,
        "store_id": store.id,
        "store_name": store.name,
        "distance_m": result["distance_m"],
        "geofence_radius_m": store.geofence_radius_m,
    }), 200

@app.post("/mobile/clock-in")
def mobile_clock_in():
    return jsonify({
        "ok": False,
        "error": "deprecated_endpoint",
        "message": "Use /api/mobile/clock-in instead."
    }), 410

# -----------------------------
# Employee Clock Page
# -----------------------------
@app.get("/employee")
def employee_page():
    stores = Store.query.order_by(Store.name.asc()).all()
    stores_min = [{"name": s.name, "code": s.qr_token} for s in stores]
    return render_template("employee_clock.html", stores=stores_min)

# -----------------------------
# Employee API (Clock In/Out)
# -----------------------------
@app.post("/api/clockin")
def api_clockin():
    process_expired_pending_auto_exits_best_effort("web_clockin")
    data = request.get_json(force=True, silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    qr_token = normalize_store_code((data.get("qr_token") or "").strip())
    lat = data.get("lat")
    lng = data.get("lng")
    accuracy_m = data.get("accuracy_m")
    allow_outside_geofence, bool_error = _json_bool_field(data, "allow_outside_geofence")
    if bool_error:
        return jsonify({"ok": False, "error": bool_error}), 400

    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin or not qr_token:
        return jsonify({"error": "Missing PIN or store code."}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"error": "Invalid or inactive employee."}), 403

    store = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
    if not store:
        log_event("CLOCKIN_DENY_INVALID_STORE", employee_pin=pin, store_code=qr_token)
        return jsonify({"error": "Invalid store code."}), 404

    open_shift = Shift.query.filter_by(employee_id=emp.id, clock_out=None).order_by(Shift.clock_in.desc()).first()
    if open_shift:
        log_event("CLOCKIN_DENY_ALREADY_CLOCKED_IN", employee_id=emp.id, open_shift_id=open_shift.id)
        return jsonify({"error": "You are already clocked in. Please clock out first."}), 409

    if device_uuid:
        other = _device_has_other_open_shift(device_uuid, emp.id)
        if other:
            log_event(
                "CLOCKIN_DENY_DEVICE_IN_USE",
                device_uuid=device_uuid,
                employee_id=emp.id,
                other_employee_id=other.employee_id,
                other_shift_id=other.id
            )
            return jsonify({"error": "This phone is currently being used for another active shift. Use your own phone or have a manager help."}), 409

    if lat is None or lng is None:
        log_event("CLOCKIN_DENY_LOCATION_REQUIRED", employee_id=emp.id, store_id=store.id)
        return jsonify({"error": "Location required."}), 400

    try:
        lat = float(lat)
        lng = float(lng)
        if accuracy_m is not None:
            accuracy_m = float(accuracy_m)
    except (TypeError, ValueError):
        log_event("CLOCKIN_DENY_BAD_LATLNG", employee_id=emp.id, store_id=store.id)
        return jsonify({"error": "Invalid lat/lng."}), 400

    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        log_event(
            "CLOCKIN_DENY_ACCURACY_TOO_LOW",
            employee_id=emp.id,
            store_id=store.id,
            accuracy_m=round(accuracy_m, 1),
            device_uuid=device_uuid or ""
        )
        return jsonify({
            "error": "GPS accuracy is too low. Step outside and try again.",
            "accuracy_m": accuracy_m
        }), 403

    dist_m = haversine_m(lat, lng, store.latitude, store.longitude)

    log_event(
        "CLOCKIN_ATTEMPT",
        employee_id=emp.id,
        employee=emp.name,
        store_id=store.id,
        store=store.name,
        store_code=store.qr_token,
        dist_m=round(dist_m, 1),
        radius_m=store.geofence_radius_m,
        device_uuid=device_uuid or ""
    )

    if dist_m > store.geofence_radius_m:
        log_event(
            "CLOCKIN_DENY_OUTSIDE_RADIUS",
            employee_id=emp.id,
            store_id=store.id,
            dist_m=round(dist_m, 1),
            radius_m=store.geofence_radius_m,
            device_uuid=device_uuid or ""
        )
        return jsonify({"error": "You are not at the store location."}), 403

    _touch_employee_device(emp, device_uuid, device_label)

    s = Shift(
        employee_id=emp.id,
        store_id=store.id,
        clock_in=now_utc(),
        clock_in_lat=lat,
        clock_in_lng=lng,
        clock_in_device_uuid=device_uuid,
        closed_by_admin=False,
        admin_closed_by=None,
        admin_closed_at=None,
        admin_close_reason=None,
    )
    db.session.add(s)
    db.session.commit()

    log_event("CLOCKIN_OK", employee_id=emp.id, shift_id=s.id, store_id=store.id, device_uuid=device_uuid or "")

    return jsonify({
        "ok": True,
        "employee": emp.name,
        "message": f"Clock-in successful for {emp.name} at {store.name}.",
        "shift_id": s.id,
        "clock_in": fmt_dt(s.clock_in),
    })

@app.post("/api/clockout")
def api_clockout():
    process_expired_pending_auto_exits_best_effort("web_clockout")
    data = request.get_json(force=True, silent=True) or {}
    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    lat = data.get("lat")
    lng = data.get("lng")
    accuracy_m = data.get("accuracy_m")

    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin:
        return jsonify({"error": "Missing PIN."}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"error": "Invalid or inactive employee."}), 403

    open_shift = Shift.query.filter_by(employee_id=emp.id, clock_out=None).order_by(Shift.clock_in.desc()).first()
    if not open_shift:
        log_event("CLOCKOUT_DENY_NO_OPEN_SHIFT", employee_id=emp.id)
        return jsonify({"error": "No open shift found. You must clock in first."}), 409

    if lat is None or lng is None:
        log_event("CLOCKOUT_DENY_LOCATION_REQUIRED", employee_id=emp.id, shift_id=open_shift.id)
        return jsonify({"error": "Location required."}), 400

    try:
        lat = float(lat)
        lng = float(lng)
        if accuracy_m is not None:
            accuracy_m = float(accuracy_m)
    except (TypeError, ValueError):
        log_event("CLOCKOUT_DENY_BAD_LATLNG", employee_id=emp.id, shift_id=open_shift.id)
        return jsonify({"error": "Invalid lat/lng."}), 400

    if accuracy_m is not None and accuracy_m > AUTO_EXIT_ACCURACY_MAX_M:
        log_event(
            "CLOCKOUT_DENY_ACCURACY_TOO_LOW",
            employee_id=emp.id,
            shift_id=open_shift.id,
            accuracy_m=round(accuracy_m, 1),
            device_uuid=device_uuid or ""
        )
        return jsonify({
            "error": "GPS accuracy is too low. Step outside and try again.",
            "accuracy_m": accuracy_m
        }), 403

    store = Store.query.get(open_shift.store_id)
    dist_m = haversine_m(lat, lng, store.latitude, store.longitude)

    log_event(
        "CLOCKOUT_ATTEMPT",
        employee_id=emp.id,
        employee=emp.name,
        shift_id=open_shift.id,
        store_id=store.id,
        store=store.name,
        store_code=store.qr_token,
        dist_m=round(dist_m, 1),
        radius_m=store.geofence_radius_m,
        device_uuid=device_uuid or ""
    )

    outside_geofence = dist_m > store.geofence_radius_m
    if outside_geofence and not allow_outside_geofence:
        log_event(
            "CLOCKOUT_CONFIRM_OUTSIDE_RADIUS",
            employee_id=emp.id,
            shift_id=open_shift.id,
            store_id=store.id,
            dist_m=round(dist_m, 1),
            radius_m=store.geofence_radius_m,
            device_uuid=device_uuid or ""
        )
        return jsonify({
            "error": "outside_store_geofence",
            "message": "You appear to be outside the store location. Clock out anyway?",
            "requires_confirmation": True,
            "dist_m": round(dist_m, 1),
            "radius_m": float(store.geofence_radius_m),
            "accuracy_m": accuracy_m,
        }), 409

    _touch_employee_device(emp, device_uuid, device_label)

    old_in = open_shift.clock_in
    old_out = open_shift.clock_out

    open_shift.clock_out = now_utc()
    open_shift.clock_out_lat = lat
    open_shift.clock_out_lng = lng
    open_shift.clock_out_device_uuid = device_uuid
    open_shift.clock_out_source = "employee"

    if outside_geofence:
        audit = ShiftEditAudit(
            shift_id=open_shift.id,
            action="employee_clock_out_outside",
            editor="WEB",
            reason=(
                f"outside_geofence=true confirmation=true source=web "
                f"lat={lat} lng={lng} dist_m={round(dist_m, 1)} "
                f"radius_m={store.geofence_radius_m} "
                f"accuracy_m={round(accuracy_m, 1) if accuracy_m is not None else 'unknown'}"
            ),
            old_clock_in=old_in,
            old_clock_out=old_out,
            new_clock_in=open_shift.clock_in,
            new_clock_out=open_shift.clock_out,
        )
        db.session.add(audit)

    _cancel_pending_auto_exit(open_shift.id, "manual_clock_out")
    db.session.commit()

    mins = shift_minutes(open_shift)
    log_event("CLOCKOUT_OK", employee_id=emp.id, shift_id=open_shift.id, minutes=mins, device_uuid=device_uuid or "")

    return jsonify({
        "ok": True,
        "employee": emp.name,
        "message": f"Clock-out successful for {emp.name}.",
        "shift_id": open_shift.id,
        "clock_out": fmt_dt(open_shift.clock_out),
        "outside_geofence": outside_geofence,
        "dist_m": round(dist_m, 1),
        "radius_m": float(store.geofence_radius_m),
        "minutes": mins,
        "human": minutes_to_human(mins),
    })

@app.post("/api/clock-status")
def api_clock_status():
    process_expired_pending_auto_exits_best_effort("web_clock_status")
    data = request.get_json(force=True, silent=True) or {}
    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()

    if not pin:
        return jsonify({"ok": False, "error": "missing_pin"}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"ok": False, "error": "invalid_or_inactive_employee"}), 403

    open_shift = (
        Shift.query
        .filter(Shift.employee_id == emp.id, Shift.clock_out.is_(None))
        .order_by(Shift.clock_in.desc())
        .first()
    )

    payload = {
        "ok": True,
        "employee": employee_payload(emp),
        "open_shift": None,
    }

    if open_shift:
        store = Store.query.get(open_shift.store_id)
        payload["open_shift"] = {
            "shift_id": open_shift.id,
            "store_id": open_shift.store_id,
            "store_name": store.name if store else "",
            "store_code": store.qr_token if store else "",
            "clock_in_local": fmt_dt(open_shift.clock_in),
        }

    return jsonify(payload), 200

# 15-minute location ping endpoint
@app.post("/api/ping")
def api_ping():
    data = request.get_json(force=True, silent=True) or {}

    username_code = normalize_employee_code(data.get("username_code") or data.get("employee_code") or "")
    pin = (data.get("pin") or "").strip()
    lat = (data.get("lat"))
    lng = (data.get("lng"))

    device_uuid = _coerce_str(data.get("device_uuid") or data.get("uuid"))
    device_label = _coerce_str(data.get("device_label"))

    if not pin:
        return jsonify({"error": "Missing PIN."}), 400

    emp = find_employee_for_mobile(username_code, pin)
    if not emp or not emp.active:
        return jsonify({"error": "Invalid or inactive employee."}), 403

    open_shift = Shift.query.filter_by(employee_id=emp.id, clock_out=None).order_by(Shift.clock_in.desc()).first()
    if not open_shift:
        return jsonify({"error": "No open shift."}), 409

    if lat is None or lng is None:
        return jsonify({"error": "Location required."}), 400

    try:
        lat = float(lat)
        lng = float(lng)
    except ValueError:
        return jsonify({"error": "Invalid lat/lng."}), 400

    store = Store.query.get(open_shift.store_id)
    dist_m = haversine_m(lat, lng, store.latitude, store.longitude)
    inside = dist_m <= store.geofence_radius_m

    _touch_employee_device(emp, device_uuid, device_label)

    ping = LocationPing(
        employee_id=emp.id,
        shift_id=open_shift.id,
        store_id=store.id,
        lat=lat,
        lng=lng,
        dist_m=float(dist_m),
        inside_radius=bool(inside),
        created_at=now_utc()
    )
    db.session.add(ping)
    db.session.commit()

    log_event(
        "PING_OK",
        employee_id=emp.id,
        shift_id=open_shift.id,
        store_id=store.id,
        dist_m=round(dist_m, 1),
        inside=inside,
        device_uuid=device_uuid or ""
    )

    return jsonify({
        "ok": True,
        "shift_id": open_shift.id,
        "dist_m": round(dist_m, 1),
        "inside_radius": inside,
        "ping_at": fmt_dt(ping.created_at),
    })

# -----------------------------
# Admin Auth
# -----------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "")

        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["admin_logged_in"] = True
            session["admin_username"] = username  # ✅ store for audit trail
            return redirect(url_for("admin_dashboard"))

        flash("Invalid username or password.", "danger")

    return render_template("login.html")

@app.get("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    session.pop("admin_username", None)
    flash("Logged out.", "info")
    return redirect(url_for("admin_login"))

# -----------------------------
# Admin Pages
# -----------------------------

@app.get("/admin")
def admin_dashboard():
    guard = admin_guard()
    if guard:
        return guard
    process_expired_pending_auto_exits_best_effort("admin_dashboard")

    total_employees = Employee.query.count()
    active_employees = Employee.query.filter_by(active=True).count()
    inactive_employees = Employee.query.filter_by(active=False).count()
    stores = Store.query.count()

    last7 = now_utc() - timedelta(days=7)
    shifts_7d = Shift.query.filter(Shift.clock_in >= last7).count()

    # -----------------------------
    # Open shifts
    # -----------------------------
    open_shifts = (
        Shift.query
        .filter(Shift.clock_out.is_(None))
        .order_by(Shift.clock_in.desc())
        .all()
    )

    open_shift_count = len(open_shifts)

    open_shift_rows = []
    longest_open_shift_rows = []
    outside_geofence_rows = []

    for s in open_shifts:
        mins = int((now_utc() - s.clock_in).total_seconds() // 60) if s.clock_in else 0

        if mins >= 600:
            color = "#dc2626"   # red
        elif mins >= 480:
            color = "#d97706"   # amber
        else:
            color = "#16a34a"   # green

        row = {
            "shift_id": s.id,
            "employee_name": s.employee.name if s.employee else "Unknown",
            "store_name": s.store.name if s.store else "Unknown",
            "clock_in": s.clock_in,
            "minutes": mins,
            "human": minutes_to_human(mins),
            "status_color": color,
        }

        open_shift_rows.append(row)
        longest_open_shift_rows.append(row)

        # Most recent ping for this open shift
        last_ping = (
            LocationPing.query
            .filter(LocationPing.shift_id == s.id)
            .order_by(LocationPing.created_at.desc())
            .first()
        )

        if last_ping and not last_ping.inside_radius:
            outside_geofence_rows.append({
                "employee_name": s.employee.name if s.employee else "Unknown",
                "store_name": s.store.name if s.store else "Unknown",
                "dist_m": round(last_ping.dist_m, 1) if last_ping.dist_m is not None else None,
                "ping_at": last_ping.created_at,
                "minutes": mins,
                "human": minutes_to_human(mins),
            })

    # Sort longest shifts first
    longest_open_shift_rows = sorted(
        longest_open_shift_rows,
        key=lambda row: row["minutes"],
        reverse=True
    )

    # Sort outside-geofence rows by farthest distance first
    outside_geofence_rows = sorted(
        outside_geofence_rows,
        key=lambda row: (row["dist_m"] is None, -(row["dist_m"] or 0))
    )

    outside_geofence_count = len(outside_geofence_rows)

    # -----------------------------
    # Recent activity
    # -----------------------------
    recent_shifts = (
        Shift.query
        .order_by(Shift.clock_in.desc())
        .limit(10)
        .all()
    )

    # -----------------------------
    # Open issues
    # -----------------------------
    open_issues_count = MobileIssueReport.query.filter(
        MobileIssueReport.status == "open"
    ).count()

    return render_template(
        "admin.html",
        total_employees=total_employees,
        active_employees=active_employees,
        inactive_employees=inactive_employees,
        stores=stores,
        shifts_7d=shifts_7d,
        open_shift_count=open_shift_count,
        open_shift_rows=open_shift_rows,
        longest_open_shift_rows=longest_open_shift_rows,
        outside_geofence_rows=outside_geofence_rows,
        outside_geofence_count=outside_geofence_count,
        recent_shifts=recent_shifts,
        open_issues_count=open_issues_count,
    )

# ✅ Admin Issues List
@app.get("/admin/issues")
def admin_issues():
    guard = admin_guard()
    if guard:
        return guard

    status = (request.args.get("status") or "open").strip().lower()
    limit_raw = (request.args.get("limit") or "").strip()

    # limit: 25..500 (default 200)
    try:
        limit = int(limit_raw) if limit_raw else 200
    except ValueError:
        limit = 200
    limit = max(25, min(limit, 500))

    q = MobileIssueReport.query

    if status in ("open", "resolved", "ignored"):
        q = q.filter(MobileIssueReport.status == status)
    else:
        status = "open"
        q = q.filter(MobileIssueReport.status == status)

    issues = q.order_by(MobileIssueReport.created_at.desc()).limit(limit).all()

    return render_template(
        "admin_issues.html",
        issues=issues,
        status=status,
        limit=limit,
    )

@app.post("/admin/issues/<int:issue_id>/set-status")
def admin_issue_set_status(issue_id: int):
    guard = admin_guard()
    if guard:
        return guard

    new_status = (request.form.get("status") or "").strip().lower()
    if new_status not in ("open", "resolved", "ignored"):
        flash("Invalid status.", "error")
        return redirect(url_for("admin_issues"))

    issue = MobileIssueReport.query.get(issue_id)
    if not issue:
        flash("Issue not found.", "error")
        return redirect(url_for("admin_issues"))

    issue.status = new_status

    if new_status == "resolved":
        issue.resolved_by = admin_username()
        issue.resolved_at = now_utc()
        # optional note
        note = (request.form.get("note") or "").strip()
        if note:
            issue.resolve_note = note
    else:
        issue.resolved_by = None
        issue.resolved_at = None
        issue.resolve_note = None

    db.session.commit()
    flash(f"Issue set to {new_status}.", "success")
    return redirect(request.referrer or url_for("admin_issues"))

from datetime import datetime

@app.post("/admin/issues/<int:issue_id>/toggle")
def admin_issue_toggle(issue_id: int):
    guard = admin_guard()
    if guard:
        return guard

    issue = MobileIssueReport.query.get(issue_id)
    if not issue:
        flash("Issue not found.", "error")
        return redirect(url_for("admin_issues"))

    # Toggle: open <-> resolved
    current = (issue.status or "open").strip().lower()

    if current == "resolved":
        issue.status = "open"
        issue.resolved_by = None
        issue.resolved_at = None
        issue.resolve_note = None
        flash("Marked issue as OPEN.", "success")
    else:
        issue.status = "resolved"
        issue.resolved_by = admin_username()
        issue.resolved_at = now_utc()
        flash("Resolved issue.", "success")

    db.session.commit()

    # Keep filters when returning
    status = (request.args.get("status") or "").strip()
    q = (request.args.get("q") or "").strip()
    page = (request.args.get("page") or "").strip()

    return redirect(url_for("admin_issues", status=status or None, q=q or None, page=page or None))

# ✅ Admin GPS Ping Viewer
@app.get("/admin/pings")
def admin_pings():
    guard = admin_guard()
    if guard: return guard
    process_expired_pending_auto_exits_best_effort("admin_pings")

    start_str = (request.args.get("start") or "").strip()
    end_str = (request.args.get("end") or "").strip()
    employee_id = (request.args.get("employee_id") or "").strip()
    store_id = (request.args.get("store_id") or "").strip()
    shift_id = (request.args.get("shift_id") or "").strip()
    inside_raw = (request.args.get("inside") or "all").strip().lower()

    try:
        page = int(request.args.get("page", "1"))
    except ValueError:
        page = 1
    page = max(page, 1)

    try:
        per_page = int(request.args.get("per_page", "200"))
    except ValueError:
        per_page = 200
    per_page = max(25, min(per_page, 500))

    if not start_str or not end_str:
        today_local = now_local().date()
        default_start = today_local - timedelta(days=7)
        start_str = start_str or default_start.isoformat()
        end_str = end_str or today_local.isoformat()

    try:
        start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
        if APP_TZ:
            start_local = datetime.combine(start_date, dtime.min, tzinfo=APP_TZ)
            end_local = datetime.combine(end_date, dtime.max, tzinfo=APP_TZ)
        else:
            start_local = datetime.combine(start_date, dtime.min)
            end_local = datetime.combine(end_date, dtime.max)
        q_start, q_end = local_range_to_utc_naive(start_local, end_local)
    except ValueError:
        flash("Invalid start/end date format. Use YYYY-MM-DD.", "error")
        today_local = now_local().date()
        start_local = datetime.combine(today_local - timedelta(days=7), dtime.min, tzinfo=APP_TZ) if APP_TZ else datetime.combine(today_local - timedelta(days=7), dtime.min)
        end_local = datetime.combine(today_local, dtime.max, tzinfo=APP_TZ) if APP_TZ else datetime.combine(today_local, dtime.max)
        q_start, q_end = local_range_to_utc_naive(start_local, end_local)
        start_str = (today_local - timedelta(days=7)).isoformat()
        end_str = today_local.isoformat()

    q = (
        LocationPing.query
        .filter(LocationPing.created_at >= q_start, LocationPing.created_at <= q_end)
        .order_by(LocationPing.created_at.desc())
    )

    if employee_id:
        try:
            q = q.filter(LocationPing.employee_id == int(employee_id))
        except ValueError:
            flash("employee_id must be a number.", "error")

    if store_id:
        try:
            q = q.filter(LocationPing.store_id == int(store_id))
        except ValueError:
            flash("store_id must be a number.", "error")

    if shift_id:
        try:
            q = q.filter(LocationPing.shift_id == int(shift_id))
        except ValueError:
            flash("shift_id must be a number.", "error")

    inside = "all"
    if inside_raw in ("1", "true", "yes", "y", "inside"):
        q = q.filter(LocationPing.inside_radius.is_(True))
        inside = "1"
    elif inside_raw in ("0", "false", "no", "n", "outside"):
        q = q.filter(LocationPing.inside_radius.is_(False))
        inside = "0"

    offset = (page - 1) * per_page
    items = q.offset(offset).limit(per_page + 1).all()
    has_next = len(items) > per_page
    pings = items[:per_page]
    has_prev = page > 1

    employees = Employee.query.order_by(Employee.active.desc(), Employee.name.asc()).all()
    stores = Store.query.order_by(Store.name.asc()).all()

    try:
        total_in_view = q.count()
    except Exception:
        total_in_view = None

    return render_template(
        "admin_pings.html",
        pings=pings,
        employees=employees,
        stores=stores,
        start=start_str,
        end=end_str,
        employee_id=employee_id,
        store_id=store_id,
        shift_id=shift_id,
        inside=inside,
        page=page,
        per_page=per_page,
        has_prev=has_prev,
        has_next=has_next,
        total_in_view=total_in_view,
    )

# ✅ Admin Mobile Event Viewer
@app.get("/admin/mobile-events")
def admin_mobile_events():
    guard = admin_guard()
    if guard:
        return guard

    event_type = (request.args.get("event") or "").strip().lower()
    device_uuid = (request.args.get("device") or "").strip()
    try:
        limit = int(request.args.get("limit", "200"))
    except ValueError:
        limit = 200
    limit = max(25, min(limit, 500))

    q = MobileEvent.query

    if event_type:
        q = q.filter(func.lower(MobileEvent.event_type) == event_type)

    if device_uuid:
        q = q.filter(MobileEvent.device_uuid == device_uuid)

    events = q.order_by(MobileEvent.received_at.desc()).limit(limit).all()

    return render_template(
        "admin_mobile_events.html",
        events=events,
        limit=limit,
        event=event_type,
        device=device_uuid,
    )

from flask import render_template

@app.get("/privacy")
def privacy_policy():
    return render_template("privacy.html")

@app.get("/admin/issues/<int:issue_id>")
def admin_issue_detail(issue_id: int):
    guard = admin_guard()
    if guard:
        return guard

    issue = MobileIssueReport.query.get(issue_id)
    if not issue:
        flash("Issue not found.", "error")
        return redirect(url_for("admin_issues"))

    # Pretty payload for template
    payload_pretty = issue.payload_json
    try:
        payload_pretty = json.dumps(json.loads(issue.payload_json), indent=2, ensure_ascii=False)
    except Exception:
        pass

    return render_template(
        "admin_issue_detail.html",
        issue=issue,
        payload_pretty=payload_pretty,
    )

@app.post("/admin/issues/<int:issue_id>/resolve")
def admin_issue_resolve(issue_id: int):
    guard = admin_guard()
    if guard:
        return guard

    issue = MobileIssueReport.query.get(issue_id)
    if not issue:
        flash("Issue not found.", "error")
        return redirect(url_for("admin_issues"))

    new_status = (request.form.get("status") or "resolved").strip().lower()
    note = (request.form.get("note") or "").strip()

    if new_status not in ("open", "resolved", "ignored"):
        new_status = "resolved"

    issue.status = new_status
    issue.resolve_note = note or None
    issue.resolved_by = admin_username()
    issue.resolved_at = now_utc() if new_status in ("resolved", "ignored") else None

    db.session.commit()
    flash("Issue updated.", "success")
    return redirect(url_for("admin_issue_detail", issue_id=issue.id))

# -------- Bulk Import (stores + employees) --------
@app.route("/admin/import", methods=["GET", "POST"])
def admin_import():
    guard = admin_guard()
    if guard: return guard

    results = None

    if request.method == "POST":
        stores_file = request.files.get("stores_file")
        employees_file = request.files.get("employees_file")

        created_stores = 0
        skipped_stores = 0
        store_errors = []

        created_emps = 0
        skipped_emps = 0
        emp_errors = []

        # ---------- Import STORES ----------
        if stores_file and stores_file.filename:
            try:
                reader = csv.DictReader(TextIOWrapper(stores_file.stream, encoding="utf-8"))
                required = {"name", "qr_token", "latitude", "longitude", "geofence_radius_m"}
                missing_cols = required - set((reader.fieldnames or []))

                if missing_cols:
                    store_errors.append(f"Stores CSV missing columns: {', '.join(sorted(missing_cols))}")
                else:
                    for i, row in enumerate(reader, start=2):
                        try:
                            name = (row.get("name") or "").strip()
                            qr_token = normalize_store_code(row.get("qr_token") or "")
                            lat = row.get("latitude")
                            lng = row.get("longitude")
                            radius = row.get("geofence_radius_m") or "150"

                            if not name or not qr_token or lat is None or lng is None:
                                skipped_stores += 1
                                store_errors.append(f"Stores row {i}: missing name/code/lat/lng")
                                continue

                            lat = float(lat)
                            lng = float(lng)
                            radius = int(float(radius))

                            existing = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
                            if existing:
                                skipped_stores += 1
                                continue

                            s = Store(
                                name=name,
                                qr_token=qr_token,
                                latitude=lat,
                                longitude=lng,
                                geofence_radius_m=radius
                            )
                            db.session.add(s)
                            created_stores += 1

                        except Exception as e:
                            skipped_stores += 1
                            store_errors.append(f"Stores row {i}: {e}")

                    db.session.commit()

            except Exception as e:
                store_errors.append(str(e))

        # ---------- Import EMPLOYEES ----------
        if employees_file and employees_file.filename:
            try:
                reader = csv.DictReader(TextIOWrapper(employees_file.stream, encoding="utf-8"))
                required = {"name", "pin"}
                missing_cols = required - set((reader.fieldnames or []))

                if missing_cols:
                    emp_errors.append(f"Employees CSV missing columns: {', '.join(sorted(missing_cols))}")
                else:
                    for i, row in enumerate(reader, start=2):
                        try:
                            name = (row.get("name") or "").strip()
                            username_code = normalize_employee_code(row.get("username_code") or row.get("employee_code") or "")
                            pin = (row.get("pin") or "").strip()
                            active_raw = (row.get("active") or "1").strip().lower()

                            if not name or not pin:
                                skipped_emps += 1
                                emp_errors.append(f"Employees row {i}: missing name or pin")
                                continue

                            active = active_raw not in ("0", "false", "no", "n")
                            if not username_code:
                                username_code = unique_employee_code_from_name(name)

                            if Employee.query.filter_by(pin=pin).first():
                                skipped_emps += 1
                                continue

                            if employee_code_exists(username_code):
                                skipped_emps += 1
                                emp_errors.append(f"Employees row {i}: username/code already in use")
                                continue

                            e = Employee(name=name, username_code=username_code, pin=pin, active=active)
                            db.session.add(e)
                            created_emps += 1

                        except Exception as e:
                            skipped_emps += 1
                            emp_errors.append(f"Employees row {i}: {e}")

                    db.session.commit()

            except Exception as e:
                emp_errors.append(str(e))

        results = {
            "created_stores": created_stores,
            "skipped_stores": skipped_stores,
            "store_errors": store_errors[:50],
            "created_emps": created_emps,
            "skipped_emps": skipped_emps,
            "emp_errors": emp_errors[:50],
        }

        flash(
            f"Import done. Stores: +{created_stores} (skipped {skipped_stores}). "
            f"Employees: +{created_emps} (skipped {skipped_emps}).",
            "success"
        )

        log_event(
            "ADMIN_IMPORT",
            created_stores=created_stores,
            skipped_stores=skipped_stores,
            created_employees=created_emps,
            skipped_employees=skipped_emps
        )

    return render_template("import.html", results=results)

@app.route("/admin/employees", methods=["GET", "POST"])
def admin_employees():
    guard = admin_guard()
    if guard:
        return guard

    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            name = (request.form.get("name") or "").strip()
            username_code = normalize_employee_code(request.form.get("username_code") or "")
            pin = (request.form.get("pin") or "").strip()

            if not name or not pin:
                flash("Name and PIN required.", "error")
            else:
                if not username_code:
                    username_code = unique_employee_code_from_name(name)

                if employee_code_exists(username_code):
                    flash("Username/code already in use.", "error")
                elif Employee.query.filter_by(pin=pin).first():
                    flash("PIN already in use.", "error")
                else:
                    e = Employee(name=name, username_code=username_code, pin=pin, active=True)
                    db.session.add(e)
                    db.session.commit()
                    flash("Employee created.", "success")

        elif action == "toggle_active":
            emp_id = request.form.get("employee_id")
            emp = Employee.query.get(emp_id)
            if emp:
                emp.active = not emp.active
                db.session.commit()
                flash(f"Employee {'activated' if emp.active else 'deactivated'}.", "success")

    view = (request.args.get("view") or "active").strip().lower()
    search_q = (request.args.get("q") or "").strip()
    store_id_raw = (request.args.get("store_id") or "").strip()
    sort_by = (request.args.get("sort") or "last_name").strip().lower()

    q = Employee.query

    if view == "inactive":
        q = q.filter(Employee.active.is_(False))
    elif view == "all":
        pass
    else:
        q = q.filter(Employee.active.is_(True))
        view = "active"

    if search_q:
        like = f"%{search_q.lower()}%"
        q = q.filter(
            (func.lower(Employee.name).like(like)) |
            (func.lower(Employee.username_code).like(like))
        )

    employees = q.all()

    stores = Store.query.order_by(Store.name.asc()).all()

    selected_store_id = None
    if store_id_raw:
        try:
            selected_store_id = int(store_id_raw)
        except ValueError:
            selected_store_id = None

    # Build helper data per employee
    employee_rows = []
    for emp in employees:
        last_shift = (
            Shift.query
            .filter(Shift.employee_id == emp.id)
            .order_by(Shift.clock_in.desc())
            .first()
        )

        last_store_id = last_shift.store_id if last_shift else None
        last_store_name = last_shift.store.name if last_shift and last_shift.store else ""
        last_clock_in = last_shift.clock_in if last_shift else None

        employee_rows.append({
            "employee": emp,
            "last_shift": last_shift,
            "last_store_id": last_store_id,
            "last_store_name": last_store_name,
            "last_clock_in": last_clock_in,
        })

    # Filter by store using last shift's store
    if selected_store_id:
        employee_rows = [
            row for row in employee_rows
            if row["last_store_id"] == selected_store_id
        ]

    # Sort
    if sort_by == "last_clock_in":
        employee_rows = sorted(
            employee_rows,
            key=lambda row: (
                row["last_clock_in"] is None,
                -(row["last_clock_in"].timestamp()) if row["last_clock_in"] else 0
            )
        )
    else:
        sort_by = "last_name"
        employee_rows = sorted(
            employee_rows,
            key=lambda row: (
                row["employee"].name.split()[-1].lower(),
                row["employee"].name.lower()
            )
        )

    inactive_count = Employee.query.filter(Employee.active.is_(False)).count()

    return render_template(
        "employees.html",
        employee_rows=employee_rows,
        employees=[row["employee"] for row in employee_rows],  # backwards compatibility if needed
        stores=stores,
        view=view,
        inactive_count=inactive_count,
        q=search_q,
        store_id=store_id_raw,
        sort=sort_by
    )

@app.post("/admin/employees/update")
def admin_employees_update():
    guard = admin_guard()
    if guard: return guard

    emp_id = request.form.get("employee_id")
    name = (request.form.get("name") or "").strip()
    username_code = normalize_employee_code(request.form.get("username_code") or "")
    pin = (request.form.get("pin") or "").strip()
    active = (request.form.get("active") or "0") == "1"

    emp = Employee.query.get(emp_id)
    if not emp:
        flash("Employee not found.", "error")
        return redirect(url_for("admin_employees"))

    if not name or not username_code or not pin:
        flash("Name, username/code, and PIN required.", "error")
        return redirect(url_for("admin_employees"))

    other_code = Employee.query.filter(
        func.lower(Employee.username_code) == username_code,
        Employee.id != emp.id
    ).first()
    if other_code:
        flash("That username/code is already in use.", "error")
        return redirect(url_for("admin_employees"))

    other = Employee.query.filter(Employee.pin == pin, Employee.id != emp.id).first()
    if other:
        flash("That PIN is already in use.", "error")
        return redirect(url_for("admin_employees"))

    emp.name = name
    emp.username_code = username_code
    emp.pin = pin
    emp.active = active
    db.session.commit()

    flash("Employee updated.", "success")
    return redirect(url_for("admin_employees"))

@app.post("/admin/employees/delete")
def admin_employees_delete():
    guard = admin_guard()
    if guard: return guard

    emp_id = request.form.get("employee_id")
    emp = Employee.query.get(emp_id)
    if not emp:
        flash("Employee not found.", "error")
        return redirect(url_for("admin_employees"))

    shift_count = Shift.query.filter_by(employee_id=emp.id).count()
    if shift_count > 0:
        flash("Cannot delete employee with shift history. Deactivate instead.", "error")
        return redirect(url_for("admin_employees"))

    db.session.delete(emp)
    db.session.commit()
    flash("Employee deleted.", "success")
    return redirect(url_for("admin_employees"))

@app.route("/admin/stores", methods=["GET", "POST"])
def admin_stores():
    guard = admin_guard()
    if guard: return guard

    if request.method == "POST":
        action = request.form.get("action")

        if action == "create":
            name = (request.form.get("name") or "").strip()
            qr_token = normalize_store_code(request.form.get("qr_token") or "")
            lat = request.form.get("latitude")
            lng = request.form.get("longitude")
            radius = request.form.get("geofence_radius_m") or "150"

            if not name or not qr_token or not lat or not lng:
                flash("Name, store code, latitude, and longitude required.", "error")
            else:
                try:
                    lat = float(lat)
                    lng = float(lng)
                    radius = int(float(radius))
                except ValueError:
                    flash("Invalid lat/lng/radius.", "error")
                else:
                    existing = Store.query.filter(func.lower(Store.qr_token) == qr_token).first()
                    if existing:
                        flash("Store code already in use.", "error")
                    else:
                        s = Store(
                            name=name,
                            qr_token=qr_token,
                            latitude=lat,
                            longitude=lng,
                            geofence_radius_m=radius
                        )
                        db.session.add(s)
                        db.session.commit()
                        flash("Store created.", "success")

    stores = Store.query.order_by(Store.name.asc()).all()
    return render_template("stores.html", stores=stores)

@app.post("/admin/stores/update")
def admin_stores_update():
    guard = admin_guard()
    if guard: return guard

    store_id = request.form.get("store_id")
    name = (request.form.get("name") or "").strip()
    qr_token = normalize_store_code(request.form.get("qr_token") or "")
    lat = request.form.get("latitude")
    lng = request.form.get("longitude")
    radius = request.form.get("geofence_radius_m") or "150"

    store = Store.query.get(store_id)
    if not store:
        flash("Store not found.", "error")
        return redirect(url_for("admin_stores"))

    if not name or not qr_token or not lat or not lng:
        flash("Name, store code, latitude, and longitude required.", "error")
        return redirect(url_for("admin_stores"))

    try:
        lat = float(lat)
        lng = float(lng)
        radius = int(float(radius))
    except ValueError:
        flash("Invalid lat/lng/radius.", "error")
        return redirect(url_for("admin_stores"))

    existing = Store.query.filter(func.lower(Store.qr_token) == qr_token, Store.id != store.id).first()
    if existing:
        flash("Store code already in use.", "error")
        return redirect(url_for("admin_stores"))

    store.name = name
    store.qr_token = qr_token
    store.latitude = lat
    store.longitude = lng
    store.geofence_radius_m = radius
    db.session.commit()

    flash("Store updated.", "success")
    return redirect(url_for("admin_stores"))

@app.post("/admin/stores/delete")
def admin_stores_delete():
    guard = admin_guard()
    if guard: return guard

    store_id = request.form.get("store_id")
    store = Store.query.get(store_id)
    if not store:
        flash("Store not found.", "error")
        return redirect(url_for("admin_stores"))

    shift_count = Shift.query.filter_by(store_id=store.id).count()
    if shift_count > 0:
        flash("Cannot delete store with shift history.", "error")
        return redirect(url_for("admin_stores"))

    db.session.delete(store)
    db.session.commit()
    flash("Store deleted.", "success")
    return redirect(url_for("admin_stores"))

@app.get("/admin/shifts")
def admin_shifts():
    guard = admin_guard()
    if guard: return guard
    process_expired_pending_auto_exits_best_effort("admin_shifts")

    shifts = Shift.query.order_by(
        Shift.clock_out.is_(None).desc(),
        Shift.clock_in.desc()
    ).limit(300).all()
    shift_ids = [s.id for s in shifts]
    pending_rows = []
    if shift_ids:
        pending_rows = (
            PendingAutoExit.query
            .filter(PendingAutoExit.shift_id.in_(shift_ids), PendingAutoExit.status == "active")
            .order_by(PendingAutoExit.deadline_at.asc())
            .all()
        )
    pending_exits_by_shift = {}
    for pending in pending_rows:
        pending_exits_by_shift.setdefault(pending.shift_id, pending)

    return render_template("shifts.html", shifts=shifts, pending_exits_by_shift=pending_exits_by_shift)

@app.post("/admin/shifts/close")
def admin_close_shift():
    guard = admin_guard()
    if guard: return guard

    shift_id = request.form.get("shift_id")
    s = Shift.query.get(shift_id)
    if not s:
        flash("Shift not found.", "error")
        return redirect(url_for("admin_shifts"))

    if s.clock_out:
        flash("Shift already closed.", "success")
        return redirect(url_for("admin_shifts"))

    old_in = s.clock_in
    old_out = s.clock_out

    s.clock_out = now_utc()
    s.clock_out_source = "admin"
    s.closed_by_admin = True
    s.admin_closed_by = admin_username()
    s.admin_closed_at = now_utc()
    s.admin_close_reason = "Admin close from shifts list"

    audit = ShiftEditAudit(
        shift_id=s.id,
        action="admin_close",
        editor=admin_username(),
        reason=s.admin_close_reason,
        old_clock_in=old_in,
        old_clock_out=old_out,
        new_clock_in=s.clock_in,
        new_clock_out=s.clock_out
    )
    db.session.add(audit)
    db.session.commit()
    flash("Shift closed.", "success")
    return redirect(url_for("admin_shifts"))

@app.post("/admin/shifts/force_close")
def admin_force_close_shift():
    guard = admin_guard()
    if guard: return guard

    shift_id = request.form.get("shift_id")
    reason = (request.form.get("reason") or "").strip()

    s = Shift.query.get(shift_id)
    if not s:
        flash("Shift not found.", "error")
        return redirect(url_for("admin_shifts"))

    if s.clock_out:
        flash("Shift already closed.", "success")
        return redirect(url_for("admin_shifts"))

    old_in = s.clock_in
    old_out = s.clock_out

    s.clock_out = now_utc()
    s.clock_out_source = "admin"
    s.closed_by_admin = True
    s.admin_closed_by = admin_username()
    s.admin_closed_at = now_utc()
    s.admin_close_reason = reason or None

    audit = ShiftEditAudit(
        shift_id=s.id,
        action="force_close",
        editor=admin_username(),
        reason=reason or "Force close (no reason provided)",
        old_clock_in=old_in,
        old_clock_out=old_out,
        new_clock_in=s.clock_in,
        new_clock_out=s.clock_out
    )
    db.session.add(audit)
    db.session.commit()

    flash("Shift force-closed (admin override).", "success")
    return redirect(url_for("admin_shifts"))

@app.route("/admin/shifts/new", methods=["GET", "POST"])
def admin_shift_new():
    guard = admin_guard()
    if guard: return guard

    employees = Employee.query.order_by(Employee.active.desc(), Employee.name.asc()).all()
    stores = Store.query.order_by(Store.name.asc()).all()

    if request.method == "POST":
        employee_id = request.form.get("employee_id")
        store_id = request.form.get("store_id")
        clock_in_raw = request.form.get("clock_in")
        clock_out_raw = request.form.get("clock_out")
        reason = (request.form.get("reason") or "").strip()

        if not employee_id or not store_id:
            flash("Employee and store are required.", "error")
            return render_template("admin_shift_new.html", employees=employees, stores=stores)

        if not reason:
            flash("Reason is required for manual shift creation.", "error")
            return render_template("admin_shift_new.html", employees=employees, stores=stores)

        cin = parse_local_datetime(clock_in_raw)
        cout = parse_local_datetime(clock_out_raw) if clock_out_raw else None

        if not cin:
            flash("Clock-in is required and must be valid.", "error")
            return render_template("admin_shift_new.html", employees=employees, stores=stores)

        if cout and cout <= cin:
            flash("Clock-out must be after clock-in.", "error")
            return render_template("admin_shift_new.html", employees=employees, stores=stores)

        s = Shift(
            employee_id=int(employee_id),
            store_id=int(store_id),
            clock_in=cin,
            clock_out=cout,
            clock_out_source="admin" if cout else None,
            closed_by_admin=True,
            admin_closed_by=admin_username(),
            admin_closed_at=now_utc(),
            admin_close_reason=reason
        )
        db.session.add(s)
        db.session.commit()

        audit = ShiftEditAudit(
            shift_id=s.id,
            action="create",
            editor=admin_username(),
            reason=reason,
            old_clock_in=None,
            old_clock_out=None,
            new_clock_in=s.clock_in,
            new_clock_out=s.clock_out
        )
        db.session.add(audit)
        db.session.commit()

        flash("Manual shift created.", "success")
        return redirect(url_for("admin_shifts"))

    return render_template("admin_shift_new.html", employees=employees, stores=stores)

@app.route("/admin/shifts/<int:shift_id>/edit", methods=["GET", "POST"])
def admin_shift_edit(shift_id: int):
    guard = admin_guard()
    if guard: return guard

    s = Shift.query.get(shift_id)
    if not s:
        flash("Shift not found.", "error")
        return redirect(url_for("admin_shifts"))

    employees = Employee.query.order_by(Employee.active.desc(), Employee.name.asc()).all()
    stores = Store.query.order_by(Store.name.asc()).all()

    if request.method == "POST":
        employee_id = request.form.get("employee_id")
        store_id = request.form.get("store_id")
        clock_in_raw = request.form.get("clock_in")
        clock_out_raw = request.form.get("clock_out")
        reason = (request.form.get("reason") or "").strip()

        if not reason:
            flash("Reason is required for shift edits.", "error")
            return render_template("admin_shift_edit.html", s=s, employees=employees, stores=stores)

        cin = parse_local_datetime(clock_in_raw)
        cout = parse_local_datetime(clock_out_raw) if clock_out_raw else None

        if not cin:
            flash("Clock-in must be valid.", "error")
            return render_template("admin_shift_edit.html", s=s, employees=employees, stores=stores)

        if cout and cout <= cin:
            flash("Clock-out must be after clock-in.", "error")
            return render_template("admin_shift_edit.html", s=s, employees=employees, stores=stores)

        old_in = s.clock_in
        old_out = s.clock_out

        s.employee_id = int(employee_id) if employee_id else s.employee_id
        s.store_id = int(store_id) if store_id else s.store_id
        s.clock_in = cin
        s.clock_out = cout
        if cout:
            s.clock_out_source = "admin"
        else:
            s.clock_out_source = None

        s.closed_by_admin = True
        s.admin_closed_by = admin_username()
        s.admin_closed_at = now_utc()
        s.admin_close_reason = reason

        audit = ShiftEditAudit(
            shift_id=s.id,
            action="edit",
            editor=admin_username(),
            reason=reason,
            old_clock_in=old_in,
            old_clock_out=old_out,
            new_clock_in=s.clock_in,
            new_clock_out=s.clock_out
        )
        db.session.add(audit)
        db.session.commit()

        flash("Shift updated (audit logged).", "success")
        return redirect(url_for("admin_shifts"))

    return render_template("admin_shift_edit.html", s=s, employees=employees, stores=stores)

@app.get("/admin/audit")
def admin_audit():
    guard = admin_guard()
    if guard: return guard

    audits = ShiftEditAudit.query.order_by(ShiftEditAudit.created_at.desc()).limit(500).all()
    return render_template("admin_audit.html", audits=audits)

# -----------------------------
# ✅ Payroll (unchanged)
# -----------------------------
@app.get("/admin/payroll")
def admin_payroll():
    guard = admin_guard()
    if guard: return guard
    process_expired_pending_auto_exits_best_effort("admin_payroll")

    start_str = request.args.get("start")
    end_str = request.args.get("end")
    out_format = (request.args.get("format") or "").lower()

    if start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
            if APP_TZ:
                start_dt = datetime.combine(start_date, dtime.min, tzinfo=APP_TZ)
                end_dt = datetime.combine(end_date, dtime.max, tzinfo=APP_TZ)
            else:
                start_dt = datetime.combine(start_date, dtime.min)
                end_dt = datetime.combine(end_date, dtime.max)
        except ValueError:
            flash("Invalid start/end date format. Use YYYY-MM-DD.", "error")
            start_dt, end_dt = last_completed_payroll_week()
    else:
        start_dt, end_dt = last_completed_payroll_week()

    q_start, q_end = local_range_to_utc_naive(start_dt, end_dt)

    shifts = Shift.query.filter(
        Shift.clock_out.isnot(None),
        Shift.clock_out >= q_start,
        Shift.clock_out <= q_end
    ).order_by(Shift.clock_out.asc()).all()

    rows = []
    totals_by_emp_min = {}
    weekly_map: dict[str, dict[int, dict[str, int]]] = {}

    for s in shifts:
        mins = shift_minutes(s)
        emp_name = s.employee.name
        store_name = s.store.name

        rows.append({
            "employee": emp_name,
            "store": store_name,
            "clock_in": fmt_dt(s.clock_in),
            "clock_out": fmt_dt(s.clock_out),
            "clock_out_source": shift_clock_out_label(s),
            "minutes": mins,
            "human_short": minutes_to_short(mins),
        })
        totals_by_emp_min[emp_name] = totals_by_emp_min.get(emp_name, 0) + mins

        cin_local = utc_naive_to_local(s.clock_in)
        wd = cin_local.weekday()  # Mon=0 ... Sun=6

        if emp_name not in weekly_map:
            weekly_map[emp_name] = {}
        if wd not in weekly_map[emp_name]:
            weekly_map[emp_name][wd] = {}
        weekly_map[emp_name][wd][store_name] = weekly_map[emp_name][wd].get(store_name, 0) + mins

    summary = []
    for emp_name in sorted(totals_by_emp_min.keys(), key=lambda x: x.lower()):
        m = totals_by_emp_min[emp_name]
        summary.append({
            "employee": emp_name,
            "minutes": m,
            "human": minutes_to_human(m),
            "human_short": minutes_to_short(m),
            "hours_decimal": minutes_to_decimal_hours(m, places=4),
        })

    grand_minutes = sum(totals_by_emp_min.values())
    grand_human = minutes_to_human(grand_minutes)
    grand_human_short = minutes_to_short(grand_minutes)
    grand_hours_decimal = minutes_to_decimal_hours(grand_minutes, places=4)

    if out_format == "csv":
        from io import StringIO

        si = StringIO()
        w = csv.writer(si)

        w.writerow(["Payroll Week Start (local)", start_dt.date().isoformat()])
        w.writerow(["Payroll Week End (local)", end_dt.date().isoformat()])
        w.writerow(["Note", "Weekly filter uses CLOCK-OUT date; day columns assign time to CLOCK-IN day (local)."])
        w.writerow([])

        day_headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        w.writerow(["Employee"] + day_headers + ["Total"])

        for emp_name in sorted(weekly_map.keys(), key=lambda x: x.lower()):
            day_cells = []
            total_emp = 0

            for wd in range(7):
                stores_for_day = weekly_map.get(emp_name, {}).get(wd, {})
                if not stores_for_day:
                    day_cells.append("0h 00m")
                    continue

                parts = []
                for store_name in sorted(stores_for_day.keys(), key=lambda x: x.lower()):
                    m = stores_for_day[store_name]
                    total_emp += m
                    parts.append(f"{store_name} {minutes_to_short(m)}")

                day_cells.append("; ".join(parts))

            w.writerow([emp_name] + day_cells + [minutes_to_short(total_emp)])

        w.writerow(["GRAND TOTAL"] + [""] * 7 + [grand_human_short])
        w.writerow([])

        w.writerow(["Shift Detail"])
        w.writerow(["Employee", "Store", "Clock In", "Clock Out", "Closed By", "Minutes", "Time (Short)"])
        for r in rows:
            w.writerow([r["employee"], r["store"], r["clock_in"], r["clock_out"], r["clock_out_source"], r["minutes"], r["human_short"]])

        output = si.getvalue()
        filename = f"payroll_{start_dt.date().isoformat()}_to_{end_dt.date().isoformat()}.csv"
        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    if out_format == "xlsx":
        from io import BytesIO

        wb = Workbook()
        logo_path = os.path.join(app.root_path, "static", "img", "company-logo.png")
        thin_gray = Side(style="thin", color="D5DCE5")
        medium_gray = Side(style="medium", color="AAB4C3")
        border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
        total_border = Border(left=thin_gray, right=thin_gray, top=medium_gray, bottom=medium_gray)
        title_font = Font(bold=True, size=18, color="1F2A44")
        subtitle_font = Font(bold=True, size=11, color="334155")
        header_font = Font(bold=True, color="1F2937")
        small_header_font = Font(bold=True, size=10, color="475569")
        total_font = Font(bold=True, color="1F2937")
        normal_font = Font(size=10, color="111827")
        warning_font = Font(bold=True, color="7F1D1D")
        wrap_top = Alignment(wrap_text=True, vertical="top")
        wrap_center = Alignment(wrap_text=True, vertical="center")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)
        total_fill = PatternFill("solid", fgColor="E8F0FE")
        header_fill = PatternFill("solid", fgColor="F1F5F9")
        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        warning_fill = PatternFill("solid", fgColor="FDE2E1")
        grand_fill = PatternFill("solid", fgColor="E2E8F0")

        def _name_parts(name: str) -> tuple[str, str, str]:
            clean = " ".join(str(name or "").split())
            if not clean:
                return "", "", ""
            parts = clean.split(" ")
            if len(parts) == 1:
                return parts[0], "", parts[0].lower()
            suffixes = {"jr", "jr.", "sr", "sr.", "ii", "iii", "iv", "v"}
            suffix = parts[-1] if parts[-1].lower().rstrip(".") in suffixes else ""
            core = parts[:-1] if suffix else parts
            surname_particles = {"da", "de", "del", "della", "der", "di", "du", "la", "le", "van", "von"}
            last_start = max(len(core) - 1, 0)
            while last_start > 0 and core[last_start - 1].lower() in surname_particles:
                last_start -= 1
            first = " ".join(core[:last_start]) if last_start > 0 else core[0]
            last = " ".join(core[last_start:]) if last_start > 0 else ""
            if suffix and last:
                last = f"{last} {suffix}"
            display = f"{last}, {first}" if last else clean
            if suffix:
                display = display if last else clean
            sort_last = (last or clean).lower()
            return display, first.lower(), sort_last

        def _format_time(dt: datetime | None) -> str:
            if not dt:
                return "Missing"
            local_dt = utc_naive_to_local(dt)
            return local_dt.strftime("%I:%M %p").lstrip("0")

        def _format_report_date(dt: datetime) -> str:
            return dt.strftime("%B %d, %Y").replace(" 0", " ")

        def _shift_hours_text(minutes: int) -> str:
            return f"{minutes_to_decimal_hours(minutes, places=2)} hrs"

        def _row_height(texts: list[str]) -> float:
            line_count = 1
            for text in texts:
                if text:
                    line_count = max(line_count, str(text).count("\n") + 1)
            return max(28, min(160, line_count * 15))

        def _shift_warnings(shift: "Shift", employee_day_shifts: list["Shift"]) -> list[str]:
            warnings = []
            if not shift.clock_in:
                warnings.append("MISSING CLOCK-IN")
            if not shift.clock_out:
                warnings.append("MISSING CLOCK-OUT")
            if not shift.clock_in or not shift.clock_out:
                warnings.append("REVIEW: INCOMPLETE TIMESTAMPS")
                return warnings
            seconds = (shift.clock_out - shift.clock_in).total_seconds()
            if seconds < 0:
                warnings.append("REVIEW: CLOCK-OUT BEFORE CLOCK-IN")
            if seconds > 16 * 60 * 60:
                warnings.append("REVIEW: SHIFT OVER 16 HOURS")
            for other in employee_day_shifts:
                if other.id == shift.id or not other.clock_in or not other.clock_out:
                    continue
                if shift.clock_in < other.clock_out and other.clock_in < shift.clock_out:
                    warnings.append("REVIEW: OVERLAPPING SHIFTS")
                    break
            return warnings

        def _excel_width_to_pixels(width: float) -> int:
            if width <= 0:
                return 0
            return int(width * 7 + 5)

        def _place_title_logo(ws, logo_path: str, first_col: int, last_col: int, row_idx: int) -> bool:
            if not os.path.exists(logo_path):
                return False
            try:
                logo = XLImage(logo_path)
                if not logo.width or not logo.height:
                    return False
                target_height_px = 32
                ratio = logo.width / logo.height
                logo.height = target_height_px
                logo.width = int(target_height_px * ratio)
                title_text_px = 245
                gap_px = 10
                title_area_px = sum(
                    _excel_width_to_pixels(ws.column_dimensions[get_column_letter(col_idx)].width or 8.43)
                    for col_idx in range(first_col, last_col + 1)
                )
                unit_width_px = logo.width + gap_px + title_text_px
                logo_left_px = max(0, int((title_area_px - unit_width_px) / 2))
                logo_top_px = 4
                remaining_px = logo_left_px
                anchor_col = first_col - 1
                for col_idx in range(first_col, last_col + 1):
                    col_px = _excel_width_to_pixels(ws.column_dimensions[get_column_letter(col_idx)].width or 8.43)
                    if remaining_px < col_px:
                        anchor_col = col_idx - 1
                        break
                    remaining_px -= col_px
                marker = AnchorMarker(
                    col=anchor_col,
                    colOff=pixels_to_EMU(remaining_px),
                    row=row_idx - 1,
                    rowOff=pixels_to_EMU(logo_top_px),
                )
                logo.anchor = OneCellAnchor(
                    _from=marker,
                    ext=XDRPositiveSize2D(pixels_to_EMU(logo.width), pixels_to_EMU(logo.height)),
                )
                ws.add_image(logo)
                return True
            except Exception as e:
                app.logger.info("Payroll XLSX title logo skipped: %s", e)
                return False

        week_dates = [start_dt + timedelta(days=offset) for offset in range(7)]
        day_headers = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        emp_records: dict[str, dict] = {}

        for shift in shifts:
            emp_name = shift.employee.name if shift.employee else "Unknown Employee"
            display_name, first_sort, last_sort = _name_parts(emp_name)
            if emp_name not in emp_records:
                emp_records[emp_name] = {
                    "display": display_name or emp_name,
                    "sort": (last_sort, first_sort, emp_name.lower()),
                    "days": {wd: [] for wd in range(7)},
                    "minutes": 0,
                }
            mins = shift_minutes(shift)
            emp_records[emp_name]["minutes"] += mins
            if shift.clock_in:
                wd = utc_naive_to_local(shift.clock_in).weekday()
            else:
                wd = 0
            if 0 <= wd <= 6:
                emp_records[emp_name]["days"][wd].append(shift)

        ws = wb.active
        ws.title = "Weekly"

        max_col = 9
        column_widths = {
            1: 24,
            2: 19,
            3: 19,
            4: 19,
            5: 19,
            6: 19,
            7: 19,
            8: 19,
            9: 13,
        }
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws["A1"] = "C&C Weekly Payroll"
        ws["A2"] = f"Payroll Week: {_format_report_date(start_dt)} through {_format_report_date(end_dt)}"
        ws["A1"].font = title_font
        ws["A1"].alignment = center
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = center
        ws.row_dimensions[1].height = 30
        _place_title_logo(ws, logo_path, 1, max_col, 1)

        summary_items = [
            ("Employees", len(emp_records)),
            ("Total Shifts", len(shifts)),
            ("Total Hours", minutes_to_decimal_hours(grand_minutes, places=2)),
        ]
        for idx, (label, value) in enumerate(summary_items, start=4):
            cell = ws.cell(row=4, column=idx)
            cell.value = f"{label}: {value}"
            cell.font = small_header_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        header_row = 5
        headers = ["Employee"] + [
            f"{day_headers[idx]}\n{week_dates[idx].strftime('%m/%d')}" for idx in range(7)
        ] + ["Total"]
        ws.append(headers)
        for col_idx in range(1, max_col + 1):
            c = ws.cell(row=header_row, column=col_idx)
            c.font = header_font
            c.alignment = center
            c.fill = header_fill
            c.border = border

        sorted_records = sorted(emp_records.values(), key=lambda item: item["sort"])
        data_start_row = header_row + 1
        for record_idx, record in enumerate(sorted_records):
            row_values = [record["display"]]
            row_warnings: dict[int, bool] = {}
            row_texts = [record["display"]]
            for wd in range(7):
                day_shifts = sorted(record["days"][wd], key=lambda item: item.clock_in or datetime.min)
                if not day_shifts:
                    row_values.append("")
                    row_texts.append("")
                    continue

                parts = []
                for shift in day_shifts:
                    mins = shift_minutes(shift)
                    warning_labels = _shift_warnings(shift, day_shifts)
                    store_name = shift.store.name if shift.store else "Unknown Store"
                    detail = [
                        store_name,
                        f"{_format_time(shift.clock_in)} - {_format_time(shift.clock_out)}",
                        _shift_hours_text(mins),
                    ]
                    if warning_labels:
                        detail.append("\n".join(dict.fromkeys(warning_labels)))
                        row_warnings[wd + 2] = True
                    parts.append("\n".join(detail))
                text = "\n\n".join(parts)
                row_values.append(text)
                row_texts.append(text)
            row_values.append(minutes_to_decimal_hours(record["minutes"], places=2))
            row_texts.append(str(row_values[-1]))
            ws.append(row_values)

            row_idx = ws.max_row
            ws.row_dimensions[row_idx].height = _row_height(row_texts)
            fill = alt_fill if record_idx % 2 else PatternFill(fill_type=None)
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = normal_font
                cell.alignment = left_center if col_idx == 1 else wrap_top
                cell.border = border
                if fill.fill_type:
                    cell.fill = fill
                if col_idx in row_warnings:
                    cell.fill = warning_fill
                    cell.font = warning_font
                if col_idx == max_col:
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.alignment = center

        grand_row = ws.max_row + 1
        ws.cell(row=grand_row, column=1).value = "GRAND TOTAL"
        ws.cell(row=grand_row, column=max_col).value = minutes_to_decimal_hours(grand_minutes, places=2)
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=grand_row, column=col_idx)
            cell.font = total_font
            cell.fill = grand_fill
            cell.border = total_border
            cell.alignment = center if col_idx == max_col else left_center

        ws.freeze_panes = f"A{data_start_row}"
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2
        ws.print_area = f"A1:{get_column_letter(max_col)}{grand_row}"
        ws.print_title_rows = f"1:{header_row}"
        ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
        ws.print_options.horizontalCentered = True

        ws2 = wb.create_sheet("Shift Detail")
        detail_headers = ["Employee", "Store", "Clock In", "Clock Out", "Closed By", "Minutes", "Time (Short)"]
        ws2.append(detail_headers)

        for col_idx in range(1, len(detail_headers) + 1):
            c = ws2.cell(row=1, column=col_idx)
            c.font = header_font
            c.font = header_font
            c.alignment = center
            c.fill = header_fill
            c.border = border

        for r in rows:
            ws2.append([r["employee"], r["store"], r["clock_in"], r["clock_out"], r["clock_out_source"], r["minutes"], r["human_short"]])

        max_col2 = len(detail_headers)
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=max_col2):
            for cell in row:
                cell.alignment = wrap_center
                cell.border = border
                if cell.row > 1:
                    cell.font = normal_font

        detail_widths = [24, 24, 22, 22, 18, 12, 14]
        for col_idx, width in enumerate(detail_widths, start=1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:{get_column_letter(max_col2)}{ws2.max_row}"
        ws2.sheet_view.showGridLines = False
        ws2.page_setup.orientation = "landscape"
        ws2.page_setup.paperSize = ws2.PAPERSIZE_LETTER
        ws2.page_setup.fitToWidth = 1
        ws2.page_setup.fitToHeight = 0
        ws2.sheet_properties.pageSetUpPr.fitToPage = True
        ws2.page_margins.left = 0.25
        ws2.page_margins.right = 0.25
        ws2.page_margins.top = 0.5
        ws2.page_margins.bottom = 0.5
        ws2.print_area = f"A1:{get_column_letter(max_col2)}{ws2.max_row}"
        ws2.print_title_rows = "1:1"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"payroll_{start_dt.date().isoformat()}_to_{end_dt.date().isoformat()}.xlsx"
        return Response(
            bio.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    day_headers = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    grid_rows = []

    for emp_name in sorted(weekly_map.keys(), key=lambda x: x.lower()):
        day_cells = []
        total_emp = 0

        for wd in range(7):
            stores_for_day = weekly_map.get(emp_name, {}).get(wd, {})
            if not stores_for_day:
                day_cells.append("0h 00m")
                continue

            parts = []
            for store_name in sorted(stores_for_day.keys(), key=lambda x: x.lower()):
                m = stores_for_day[store_name]
                total_emp += m
                parts.append(f"{store_name} {minutes_to_short(m)}")

            day_cells.append("; ".join(parts))

        grid_rows.append({
            "employee": emp_name,
            "days": day_cells,
            "total": minutes_to_short(total_emp),
        })

    return render_template(
        "payroll.html",
        start=start_dt.date().isoformat(),
        end=end_dt.date().isoformat(),
        summary=summary,
        rows=rows,
        day_headers=day_headers,
        grid_rows=grid_rows,
        grand_minutes=grand_minutes,
        grand_human=grand_human,
        grand_human_short=grand_human_short,
        grand_hours_decimal=grand_hours_decimal
    )

# ✅ Backwards-compatible alias for old Reports link
@app.get("/admin/reports/hours")
def admin_reports_hours_redirect():
    guard = admin_guard()
    if guard:
        return guard

    args = request.args.to_dict(flat=True)
    return redirect(url_for("admin_payroll", **args))

# -----------------------------
# Index
# -----------------------------
@app.get("/")
def index():
    return redirect(url_for("employee_page"))

# -----------------------------
# Run (local only)
# -----------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
